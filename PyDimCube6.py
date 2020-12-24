#PyDimCube
Version=6
#PyDimCube6
#Distributing the jet and pool fire frequenciencies
#
PropJetNotPool = 0.5
#For a cube, read from each cube the duration that jet fire impinges on it
#PyDimCube3
#To consider failure cases for fire or gas detector
#Failure of fire detector for immediate ignition
#Failure of gas detectin for delayed ignition
#PyDimCube4 using 'Bv06_c.xlsx'


from openpyxl import load_workbook
import math
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import scipy.ndimage
from scipy.ndimage.filters import gaussian_filter
import dill
import sys

import matplotlib.pylab as pltparam

#AdjModules
AdjModules = {}
AdjModules['S05'] = ['S04','P05','P04','KOD']
AdjModules['S04'] = ['S05','S03','P05','P04','P03']
AdjModules['S03'] = ['S04','S02','P04','P03','P02']
AdjModules['S02'] = ['S03','P03','P02']
AdjModules['P05'] = ['P04','S05','S04','KOD']
AdjModules['P04'] = ['P05','P03','S05','S04','S03']
AdjModules['P03'] = ['P04','P02','S04','S03','S02']
AdjModules['P02'] = ['P03','S03','S02']
AdjModules['R02'] = ['P03','S03','P02','S02']
AdjModules['R03'] = ['P04','S04','P03','S03','P02','S02']
AdjModules['R04'] = ['P04','S04','P03','S03','P05','S05']
AdjModules['R05'] = ['P04','S04','P05','S05']
AdjModules['KOD'] = ['P05','S05']
AdjModules['TUR'] = ['KOD','P05','S05']
AdjModules['FWD'] = ['KOD','P05','S05']
#HullDeck
AdjModules['HD_AP'] = ['HD_FP','HD_FS','HD_AP','HD_AS']
AdjModules['HD_AS'] = ['HD_FP','HD_FS','HD_AP','HD_AS']
AdjModules['HD_FP'] = ['HD_FP','HD_FS','HD_AP','HD_AS']
AdjModules['HD_FS'] = ['HD_FP','HD_FS','HD_AP','HD_AS']
#Reel Station
AdjModules['LQ_A'] = ['LQ_A','LQ_P','LQ_S']
AdjModules['LQ_P'] = ['LQ_A','LQ_P','LQ_S']
AdjModules['LQ_S'] = ['LQ_A','LQ_P','LQ_S']


# for am in AdjModules:
#     AdjModules[am].append(am)

# sys.stdout = open('DimCube4.txt','w')

FireWallBtn3and4 = False
WantPlot = False
cube_result2file = True
FirewallX = 140.7 #meter from AP

DimCubeSuccessList = {}
DimCubeFailList = {}
def DtoH (D05):
    if D05 > 19.7:
            H05 = 0.4664*D05 + 18.345
    elif D05 < 1.0:
        H05 = 0.
    else: #  1.0 < D05 < 19.7:
        H05 = 1.2592*D05 + 2.7235
    return H05
def PD(t,De,Te):
    if t < Te:
        PD = De * (1-math.sqrt(t/Te))
    else:
        PD = 0
    return PD
#pool burning rate 0.062 kg/m2-s
#MS ; spilt mass
#Pool fire duration
# Ms / (3.14*D^2/4 * br) *3/2 where 3/2 is a factor to consider shringking pool

def print_cum_cube(cube,AA):
    F=0.
    print("{:50s} Mod  (Freq. ) - Jet Duration  CumFreq".format(cube))
    for e in AA[::-1]:
        # pv,hole,weather = e[2].split("\\")
        # hole = hole.split("_")[0]
        F += e[1]
        pv,hole,weather = e[2].split("\\")
        mod = Modules[pv]
        print("{:50s}{:5s} {:8.2e} {:8.1f}   {:8.2e}".format(e[2],mod,e[1],e[0],F))
        if F>1.0E-3:
            break
def print_cum_cube_file(cube,AA,a_file):
    F=0.
    print("{:50s} Mod  (Freq. ) - Jet Duration  CumFreq".format(cube),file = a_file)
    for e in AA[::-1]:
        # pv,hole,weather = e[2].split("\\")
        # hole = hole.split("_")[0]
        F += e[1]
        pv,hole,weather = e[2].split("\\")
        mod = Modules[pv]
        print("{:50s}{:5s} {:8.2e} {:8.1f}   {:8.2e}".format(e[2],mod,e[1],e[0],F),file = a_file)
        if F>1.0E-3:
            break

# iIS=load_workbook(filename='IS_v12_shk.xlsx',data_only=True)
iIS=load_workbook(filename='IS_v15_for_SHI_20201127.xlsx',data_only=True)
shIS = iIS['Isolatable summary']
IS_sub = {}
numESDVs = {}
Modules = {}
Deck = {}
IS_P = {}
IS_T = {}
IS_rho = {}
IS_V = {}
IS_Vadj = {}
r = 3
# while r < 79:
#     nsub = shIS.cell(r,4).value
#     IS_sub[shIS.cell(r,3).value] = [r,nsub]
#     r += nsub
while r < 81:
    #for IS_v12
    # pv = shIS.cell(r,5).value
    # IS_sub[pv] = shIS.cell(r,11).value #Read for each leak at respective height
    # Modules[pv] = shIS.cell(r,7).value
    # Deck[pv] =  shIS.cell(r,8).value

    # IS_P[pv] = shIS.cell(r,38).value
    # IS_T[pv] = shIS.cell(r,39).value
    # IS_rho[pv] = shIS.cell(r,40).value
    # IS_V[pv] = shIS.cell(r,32).value
    # IS_Vadj[pv] = shIS.cell(r,33).value    
    # if shIS.cell(r,24).value != None:
    #    nedvs = shIS.cell(r,24).value.count("\"")    
    #    numESDVs[pv] = nedvs
    # else:
    #    numESDVs[pv] = pnedvs
    #For IS_v15
    pv = shIS.cell(r,3).value
    IS_sub[pv] = shIS.cell(r,9).value #Read for each leak at respective height
    if IS_sub[pv] == 'NA':
        IS_sub[pv] = 1         
    Modules[pv] = shIS.cell(r,5).value
    Deck[pv] =  shIS.cell(r,6).value

    IS_P[pv] = shIS.cell(r,39).value
    IS_T[pv] = shIS.cell(r,40).value
    IS_rho[pv] = shIS.cell(r,41).value
    IS_V[pv] = shIS.cell(r,33).value
    IS_Vadj[pv] = shIS.cell(r,34).value    
    if shIS.cell(r,23).value != None:
       nedvs = shIS.cell(r,23).value.count("\"")    
       numESDVs[pv] = nedvs
    else:
       numESDVs[pv] = pnedvs
    pnedvs = nedvs
    r += 1


import pandas as pd
pdIS = pd.DataFrame([Modules, Deck, IS_P, IS_T,IS_V,IS_Vadj],index=['Module','Deck','P','T','Vol','VolAdj'])
pdIS = pdIS.transpose()
pdIS = pdIS.replace(0,np.nan)
pdIS = pdIS.fillna(method='ffill')
#
pdIS_M = pdIS.groupby('Module')
pdIS_M['P','Vol'].agg('mean')
#
id_p05 = pdIS['Module'] == 'P05'
pdIS[id_p05].mean()
# P         121.3000


import dill
Area = "ProcessArea"
element_dump_filename = 'Bv08_c5_dump'
icubeloc='SCE_CUBE_XYZ2_Process'
FractionPoolFire = 1

# Area = "HullDeck"
# element_dump_filename = 'Bv06_hull_dump'
# icubeloc='SCE_CUBE_XYZ2_HullDeck'
# FractionPoolFire = 1

# Area = "Reel Station"
# element_dump_filename = 'Bv06_offloading_dump'
# icubeloc='SCE_CUBE_XYZ2_Offloading'
# FractionPoolFire = 1

# element_dump_filename = 'Bv06_utility_dump'
# icubeloc='SCE_CUBE_XYZ2_Utility'
# FractionPoolFire = 1

with open(element_dump_filename,'rb') as element_dump:
    lEvent = dill.load(element_dump)

#Read 'Presure vessel' from Input
# iExlFilename='Bv06_i'
iExlFilename='Bv08_i5'
iExl=load_workbook(filename=iExlFilename+'.xlsx')
shPV = iExl['Pressure vessel']
X = {}
Y = {}
Z = {}
r=63
while shPV.cell(r,1).value == "Yes":    
    study  = shPV.cell(r,2).value
    pv  = shPV.cell(r,8).value    
    # key = study  + "\\" +  pv
    key = pv    
    X[key] = shPV.cell(r,136).value
    Y[key] = shPV.cell(r,137).value    
    Z[key] = shPV.cell(r,20).value        #Elevation
    r += 1
numPV = r-63
def jffit(m):
    jl_lowe = 2.8893*np.power(55.5*m,0.3728)
    if m>5:
        jf = -13.2+54.3*math.log10(m)
    elif m>0.1:
        jf= 3.736*m + 6.
    else:
        jf = 0.
    # print(m, jl_lowe,jf)
    return jl_lowe


# iExlFilename='025-02-01-G_i'
# iExl=load_workbook(filename=iExlFilename+'.xlsx')
# cExlFilename='025-02-01-G_c'

#Failure of fire detector for immediate ignition
P_FD_Fail = 0.05 #default value? #should be adjusted for hole size or leak rate? To be suggested to use the release rate!
# P_FD_Fail['SM'] = 0.1
# P_FD_Fail['ME'] = 0.05
# P_FD_Fail['MA'] = 0.005
# P_FD_Fail['LA'] = 0.005

#Failure of gas detectin for delayed ignition
P_GD_Fail = 0.05

#PyDimCube5 to distinguish release before and after the hull deck coaming
X_hd_coaming = 120.5 

iExl=load_workbook(filename=icubeloc+'.xlsx')
shCube = iExl['xyz']
ncube = shCube.cell(1,1).value
Xcube = {}
Ycube = {}
Zcube = {}
Cubes = []
CubeDeck = {}
CubeImpingeDuration = {}
for i in range(3,ncube+3):
    id = shCube.cell(i,1).value
    Cubes.append(id)
    Xcube[id] = shCube.cell(i,2).value
    Ycube[id] = shCube.cell(i,3).value
    Zcube[id] = shCube.cell(i,4).value
    CubeDeck[id] = shCube.cell(i,5).value

DrainRateModuleVol = {'S05':409.4,'S04':409.4,'S03':409.4,'S02':253.1,'P05':409.4,'P04':520,'P03':409.4,'P02':409.4,'KOD':409.4}
DrainRateModuleArea = {'S05':646.5,'S04':584.0,'S03':483.4,'S02':474.2,'P05':585,'P04':584,'P03':475,'P02':403,'KOD':403}
EquipCoamingArea = {'S05':3,'S04':33,'S03':10,'S02':0,'P05':15,'P04':8,'P03':6,'P02':40}#the smallest drain  is used
EquipCoamingVol = {'S05':138.7,'S04':147.7,'S03':138.7,'S02':0,'P05':138.7,'P04':138.7,'P03':138.7,'P02':416.1}#the smallest drain  is used
#print Header
print("Scenario                                       Module Deck   (     x,     y,     z )    Distance   Cube      (     x,     y,     z ) Dur[sec]  Dur[min]   m.[0]  m.[5min]  m.[T]")

ImpingeDurationArray = []

DimScn = {}
# for i in range(0,1):    
for i in range(0,ncube):    
    # id = Cubes[i]
    # CubeModule = id[:3]
    # for e in lEvent:
    #     if (e.Module in AdjModules[CubeModule]):
    #         print(CubeModule," affected by ", e.Module)
# for i in range(0,37):
# for i in [ncube-1]:
    # li = 0
    # sep = np.zeros((len(lEvent),2))
    ImpingeDurationArray = []
    id = Cubes[i]
    if (Area == "HullDeck") or (Area == 'Reel Station'):
        CubeModule = id
    else:
        CubeModule = id[:3]
    xx = Xcube[id]
    yy = Ycube[id]
    zz = Zcube[id]
    if cube_result2file == True:
        f_cube_result = open(id+"_"+str(Version)+".txt","w")
    # if (zz >= 35.0) and (zz < 44):
    #     CubeDeck = 'A'
    # elif (zz >= 44.0) and (zz < 53):
    #     CubeDeck = 'B'
    # elif (zz >= 53):
    #     CubeDeck = 'C'
    # elif (zz < 35 ):
    #     CubeDeck = 'Hull'
    
    for ei in range(0,len(lEvent)):
        e = lEvent[ei]
        
        dx = xx - e.X
        dy = yy - e.Y
        if e.Deck == "A":            
            DeckHeight = 35
        elif e.Deck == "B":
            DeckHeight = 44
        elif e.Deck == "C":
            DeckHeight = 53
        elif e.Deck == "Hull":
            DeckHeight = 29
        else:
            print("Wrong deck definiton")
            exit()
        dz = zz - (e.ReleaseHeight+DeckHeight)
        rr = math.sqrt(dx*dx+dy*dy+dz*dz)
        rr2 = math.sqrt(dx*dx+dy*dy)
        #to consider only the vapour fraction
        RelRate = e.Discharge.ReleaseRate*(1-e.Discharge.LiquidMassFraction+0.05)
        # rr = e.Discharge.ReleaseRate
        if RelRate <= 1:
            P_FD_Fail = 0.1
            P_GD_Fail = 0.1
        elif RelRate <= 10:
            P_FD_Fail = 0.05
            P_GD_Fail = 0.05
        elif RelRate > 10:
            P_FD_Fail = 0.005
            P_GD_Fail = 0.005
        else:
            print('something wrong in P_FD_Fail')
            exit()

                        
            
        #Pool fire
        Cond1 = ((CubeModule == e.Module) and (CubeDeck[id] == e.Deck) and ((e.EarlyPoolFire != None) or (e.LatePoolFire != None)) )
        Cond2 = (e.Module[:2]=="HD") and (e.Deck == "Hull" ) and (((id[-1] == 'A') and (e.X < X_hd_coaming)) or ((id[-1] == 'F') and (e.X >= X_hd_coaming )))
        t_pool_early = t_pool_late = 0.
        # print(e.Key, Cond1, Cond2)        
        # if ((CubeModule == e.Module) and (CubeDeck[id] == e.Deck) and ((e.EarlyPoolFire != None) or (e.LatePoolFire != None)) ) :        
        # if ((CubeDeck[id] == e.Deck) and ((e.EarlyPoolFire != None) or (e.LatePoolFire != None)) ) or (((id[-1] == 'A') and (e.X < X_hd_coaming)) or ((id[-1] == 'F') and (e.X >= X_hd_coaming ))) :        
        if  Cond1 or Cond2 :
            # print(e.Key, e.Module,"I am in pool fire")

            #Density
            #To handle the case that 'Deck' is designated 'B' or 'C'
            is_name, is_hole, is_weather = e.Key.split("\\")
            if is_name[-1] in ['A','B','C']:
                for deck_id in ['A','B','C']:
                    if is_name[:-1]+deck_id in IS_sub.keys():
                        if IS_rho[is_name[:-1]+deck_id] != None:
                            rho = IS_rho[is_name[:-1]+deck_id]
                            CondensateDensity = rho #kg/m3            
            elif '131-01' in is_name:
                CondensateDensity = 778.8 #kg/m3
            else:
                CondensateDensity = IS_rho[is_name]
            
            #Time for the release rate is larger than drain-rate
            if e.Deck == 'B':
                # drain_rate = (EquipCoamingArea[e.Module]*0.27)/3600*CondensateDensity                #0.27 m/hr (270mm/hr) "Open drain sizing Section 7 1)" for overboarding???
                # drain_rate = (EquipCoamingArea[e.Module]*DrainRateModuleVol[e.Module]/DrainRateModuleArea[e.Module])/3600*CondensateDensity                #0.27 m/hr (270mm/hr) "Open drain sizing Section 7 1)" for overboarding???
                drain_rate = EquipCoamingVol[e.Module]/3600*CondensateDensity                #0.27 m/hr (270mm/hr) "Open drain sizing Section 7 1)" for overboarding???
                drain_area_dia = math.sqrt(EquipCoamingArea[e.Module]*4/3.14)
            elif e.Deck == 'A':
                drain_rate = DrainRateModuleVol[e.Module]/3600*CondensateDensity
                drain_area_dia = math.sqrt(DrainRateModuleArea[e.Module]*4/3.14)
            else: 
                drain_rate = 0
                drain_area_dia = 100

            if e.TVD[-1,2] > drain_rate: #if the leak rate at the instant of leak stop is higher thatn drain_rate...
                t_pool_release = e.TVD[-1,0]
            elif e.TVD[1,2] > drain_rate: #if the initial leak rate is higher than the drain_rate but the last leak rate is lower than 'drain_rate'...           
                t_e = interp1d(e.TVD[:,2],e.TVD[:,0],kind='linear') #interpolated model for release rate t_e(rr)  will give the time that the release rate becomes 'rr'
                #but the release before the time would have been stacked to keep the pool fire more... Should additional time be considered?

                t_pool_release = t_e(drain_rate)
            else: #release is lower than the initial release rate
                # print(e.Key, e.TVD[1,2], e.TVD[-1,2], drain_rate)
                t_pool_release = 0.
                if e.EarlyPoolFire != None:
                    e.EarlyPoolFire.Frequency = 0.
                if e.LatePoolFire != None:
                    e.LatePoolFire.Frequency = 0.
                # print(id,is_name, is_hole,is_weather,e.Module,e.Deck," is neglected by setting t_pool_releae = 0.", CondensateDensity, drain_rate,)
            

                # To check how much pool fire events are neglected
                # for e in lEvent:
                #     if (e.EarlyPoolFire != None) and (e.EarlyPoolFire.Frequency == 0):
                #         print(e.Key,"EPF Frequency set to zero")
                #     if (e.LatePoolFire != None) and (e.LatePoolFire.Frequency == 0):
                #         print(e.Key,"LPF Frequency set to zero")


            #Time to burn mass spilt and collected in the coaming considering drain
            burn_rate = 0.062
            #No equipment coaming and drain for the module?
            Mass_drain = t_pool_release * drain_rate
            # Ms_to_burn = max(e.Discharge.Ms[0] - e.Discharge.Ms[-1] - Mass_drain,0) #if coaming is too small to contain release even considering the drain, it should be used.                
            # f_disc = interp1d(e.Discharge.Ts,e.Discharge.Ms)  
            f_disc = interp1d(e.TVD[:,0],e.TVD[:,1])  
            if t_pool_release > 300:
                MassAtPoolReleaseStop = f_disc(t_pool_release)#can't deal with the case tht releast stops before 5 min. Release quantity is overestimated to be the total mass           
            else:
                MassAtPoolReleaseStop = e.TVD[-1,1]
            Ms_to_burn = max(e.Discharge.Ms[0] - MassAtPoolReleaseStop - Mass_drain,0) #if coaming is too small to contain release even considering the drain, it should be used.                
            if e.BundArea > 0.:
                #Coaming
                # Volume                     
                CoamingArea = e.BundArea #m2
                CoamingHeight = 0.15 #m
                
                Mass_in_coaming = CoamingArea * CoamingHeight*CondensateDensity
                # if Mass_in_coaming < Ms_to_burn:
                #     print(e.Key, "Drip tray overfilled", IS_rho[is_name],Mass_in_coaming, Ms_to_burn)
                Ms_to_burn = min(Mass_in_coaming, Ms_to_burn) #if coaming is too small to contain release even considering the drain, it should be used.                        
            
            # print(id,e.Key,t_pool_release,Ms_to_burn)
            if (e.EarlyPoolFire != None) and (e.EarlyPoolFire.Frequency != 0.):
                dd = min(e.EarlyPoolFire.Diameter,drain_area_dia)
                t_pool_burn = Ms_to_burn / (3.14/4*dd**2 * burn_rate)
                
                t_pool_early = max(t_pool_release,t_pool_burn)
                
                e.EarlyPoolFire.PoolFireDurations = [t_pool_release,t_pool_burn,t_pool_early]
                

                # if (rr < dd*0.5) & (zz-35 < DtoH(dd)) & (t_pool_early > 0): # cube is within the pool, cube center is higher than the pool height, duation > 0
                # if (rr2 < dd*0.5) & (t_pool_early > 0): # cube is within the pool, cube center is higher than the pool height, duation > 0
                if (t_pool_early > 0): # cube is within the pool, cube center is higher than the pool height, duation > 0
                    # if (rr2 < 1*dd*0.5): # cube is within the pool, cube center is higher than the pool height, duation > 0
                        # print("Exposre to Early pool fire",e.Key)
                        # di = PFD*(dd-2*rr)/dd   
                        di = t_pool_early                     
                        # print('Release',t_pool_release,'Burn',t_pool_burn,'Early Pool fire duration',di)

                        if ('EXBX' in e.Key) or ('EXBN' in e.Key):
                            #Fire detection successful but ESD fail, FO_EX                            
                            ff = e.EarlyPoolFire.Frequency*(1-P_FD_Fail)*(1-PropJetNotPool)
                            ImpingeDurationArray.append([di,ff,e.Key+"FO_EarlyPool"])                        
                            #Fire detection failure and ESD was not activated, FX_EX 
                            ff = 0.
                            pv,hole,weather = e.Key.split("\\")
                            for ee in lEvent:
                                if ((pv in ee.Key) and (hole[:2] == ee.Hole[:2]) and (weather == ee.Weather)):
                                    ff += ee.EarlyPoolFire.Frequency*P_FD_Fail*(1-PropJetNotPool)
                            ImpingeDurationArray.append([di,FractionPoolFire*ff,e.Key+"FX_EarlyPool"])                                                    
                        else:                          
                            ImpingeDurationArray.append([di,FractionPoolFire*e.EarlyPoolFire.Frequency*(1-P_FD_Fail)*(1-PropJetNotPool),e.Key+"FO_EarlyPool"])
                    #     print(id, e.Key, e.Module,di,"Early pool fire added")
                    # else:
                    #     print('Early pool too small to expose Cube?',id, e.Key, e.Module, e.Deck, rr2,dd*0.5, t_pool_early)
            
            if (e.LatePoolFire != None)  and (e.LatePoolFire.Frequency != 0.):
                #Read pool fire duration, ho                    
                dd = min(e.LatePoolFire.Diameter,drain_area_dia)
                t_pool_burn = Ms_to_burn / (3.14*dd**2/4 * burn_rate)
                t_pool_late = max(t_pool_release,t_pool_burn)  
                e.LatePoolFire.Ts = [t_pool_release,t_pool_burn,t_pool_late]

                # if (rr < dd*0.5) & (zz - 35 < DtoH(dd)) & (t_pool_late > 0):
                # if (rr2 < dd*0.5) & (t_pool_late > 0):
                if (t_pool_late > 0):
                    # if (rr2 < 1*dd*0.5):
                        # print("Exposre to Late pool fire",e.Key)
                        # di = PFD*(dd-rr)/dd                        
                        di = t_pool_late                     
                        # print('Release',t_pool_release,'Burn',t_pool_burn,'Late Pool fire duration',di)
                        if ('EXBX' in e.Key) or ('EXBN' in e.Key):
                            #Fire detection successful
                            ff = e.LatePoolFire.Frequency*(1-P_GD_Fail)
                            ImpingeDurationArray.append([di,ff,e.Key+"FO_LatePool"])                        
                        #Fire detection failure
                            ff = 0.
                            pv,hole,weather = e.Key.split("\\")
                            for ee in lEvent:
                                if ((pv in ee.Key) and (hole[:2] == ee.Hole[:2]) and (weather == ee.Weather)):
                                    ff += ee.LatePoolFire.Frequency*P_GD_Fail
                            ImpingeDurationArray.append([di,FractionPoolFire*ff,e.Key+"FX_LatePool"])                                                    
                        else:                          
                            ImpingeDurationArray.append([di,FractionPoolFire*e.LatePoolFire.Frequency*(1-P_GD_Fail),e.Key+"FO_LatePool"])
                    #     print(id, e.Key, e.Module,di,"Late pool fire added")
                    # else:
                    #     print('Late pool too small to expose Cube?',id, e.Key, e.Module, e.Deck, rr2,dd*0.5, t_pool_late)                        

        #Jet Fire Analysis
        # if (e.JetFire != None): #to consider all release
        # if (e.JetFire != None) and ((CubeDeck[id] == e.Deck) or (e.Module == CubeModule)): #To consider release at the same deck or in the same module                
        # if (e.JetFire != None) and ((e.Module == CubeModule) or (e.Module in AdjModules[CubeModule])) : #To consider release sources at any decks in the adjacent modules        
        if (e.JetFire != None) and ((e.Module == CubeModule) or ((e.Module in AdjModules[CubeModule]) and (CubeDeck[id] == e.Deck))) : #To consider release sources at the same deck in the adjacent modules        
            
            #Array of jet length
            jl_e = e.jfscale*2.8893*np.power(55.5*e.TVD[:,2],0.3728) #the 3rd column of the matrix, Lowesmith formulat
            t_e = interp1d(jl_e,e.TVD[:,0],kind='linear') #Read the time when 'jl' is equal to the distance 'rr'            
            
            # The following condition is to consider effects of fire wall
            # If the firewall is considered and the source and targets are at different areas, we skip considering the jet fire effect.
            # In other words, if the fire is not considered (firewall == false) or (if they are in the same area), we will consider the jet fire
            # if (FireWallBtn3and4 == False) or (((xx > FirewallX) and (e.X > FirewallX)) or ((xx < FirewallX) and (e.X < FirewallX)))
            # if (FireWallBtn3and4 == False) or (((xx > FirewallX) and (e.X > FirewallX)) or ((xx < FirewallX) and (e.X < FirewallX))):            

            if t_pool_early > 0.:
                JetFireFrequency = PropJetNotPool*e.JetFire.Frequency
            else:
                JetFireFrequency = e.JetFire.Frequency

            #With no ESD
            if ('EXBX' in e.Key) or ('EXBN' in e.Key):
                #With EXBX or EXBN, there are two cases of ESDV fail. One is when fire detection fails and the other is when FD is successfule but ESDV itself fails
                #Fire detection successful
                #Probability of failure of ESD and BDV is already reflected in e.Frequency (Leak frequency)
                ff = JetFireFrequency*(1-P_FD_Fail)
                if max(jl_e) < rr:                                     
                    ImpingeDurationArray.append([0.,0.,e.Key+"FO_Jet"])                    
                elif min(jl_e) > rr:                
                    ImpingeDurationArray.append([e.TVD[-1,0],ff,e.Key+"FO_Jet"])                             
                else:                
                    ImpingeDurationArray.append([float(t_e(rr)),ff,e.Key+"FO_Jet"])                    

                #Fire detection failure, any release irrespective of success of ESDV and BDV will result in EXBX release case      
                ff = JetFireFrequency/e.PESD/e.PBDV*P_FD_Fail
                if max(jl_e) < rr:                                     
                    ImpingeDurationArray.append([0.,0.,e.Key+"FX_Jet"])                    
                elif min(jl_e) > rr:                
                    ImpingeDurationArray.append([e.TVD[-1,0],ff,e.Key+"FX_Jet"])                             
                else:                
                    ImpingeDurationArray.append([float(t_e(rr)),ff,e.Key+"FX_Jet"])                    
            else:  
                #Fire detection succedded
                #Probability of success of ESD is already reflected in e.Frequency
                ff = JetFireFrequency*(1-P_FD_Fail)
                if max(jl_e) < rr:                                     
                    ImpingeDurationArray.append([0.,0.,e.Key+"FO_Jet"])                    
                elif min(jl_e) > rr:                
                    ImpingeDurationArray.append([e.TVD[-1,0],ff,e.Key+"FO_Jet"])                             
                else:                
                    ImpingeDurationArray.append([float(t_e(rr)),ff,e.Key+"FO_Jet"])    


    #To pin-point a scenario that give the dimensioning scenario
    IDAsorted = sorted(ImpingeDurationArray, key = lambda fl: fl[0]) #with the longest duration at the bottom
    
    cf = 0
    jp = IDAsorted[-1]
    DimFreq = 1.0E-4
    di = 0
    cfp = jp[1]#frequency for the longest duration jet fire
    cf = cfp
    InterpolationSuccess = False
    if cfp > DimFreq:
        di = jp[0]
        scn = jp[2]
    else:
        for j in IDAsorted[-2:0:-1]:
            cf = cfp + j[1]
            scn = j[2]
            di = j[0]
            #To read the scenario that is closet to the dim frequency regarded of suceess of ESDV and BDV
            # if cf >= DimFreq and cfp < DimFreq:
            #     dp = jp[0]
            #     dn = j[0]
            #     # di = (dn-dp)/(cf-cfp)*(DimFreq - cfp) + dp
            #     # print('Dimensioning jet duration {:8.1f}'.format(di))
            #     scnp = jp[2]
            #     scn = j[2]
            #     InterpolationSuccess = True
            #     #Choose the scenario close to the threshold
            #     if abs(cf-DimFreq) > abs(cfp-DimFreq):
            #         scn = scnp
            #         di = dp
            #     else:
            #         di = dn
            #     break
            #End of "To read the scenario that is closet to the dim frequency regardeless of suceess of ESDV and BDV"
            
            if (cf >= DimFreq) and (("EOBO" in scn) or ("EOBN" in scn)):
                InterpolationSuccess = True
                break            
            cfp = cf
            # print(cfp,cf)
            jp = j
    
    # print(scn)
    if InterpolationSuccess:
        CubeImpingeDuration[Cubes[i]] = di
        if "Pool" in scn:
            for e in lEvent:
                if e.Key in scn:
                    break
            pv,hole,weather = scn.split("\\")
            weather,pl = weather.split("_")
            x1 = X[pv]
            y1 = Y[pv]
            z1 = Z[pv]
            if (e.Deck == "Hull"):
                z1 += 29
            else:
                z1 = z1+35
            
            dx=xx-x1
            dy=yy-y1
            dz=zz-z1
            ll = math.sqrt(dx*dx+dy*dy+dz*dz)
            if "Early" in pl:
                PFD = e.EarlyPoolFire.Diameter
            elif "Late" in pl:
                PFD = e.LatePoolFire.Diameter
            else:
                print("343: something wrong")
            pv,hole,weather = scn.split("\\")
            safety = hole[3:]    
            hole = hole[:2]
            weather,jet_or_pool = weather.split("_")
            fgdet = weather[-2:]
            weather = weather[:-2]
            print("{:10s} {:20s} {:2s} {:2s} {:4s} {:6s} {:5s} {:4s} ({:6.1f},{:6.1f},{:6.1f} ) -> {:6.1f} -> {:10s} ({:6.1f},{:6.1f},{:6.1f} ) {:8.1f} {:8.1f} {:8.1f}".\
                format(jet_or_pool,pv,hole,fgdet,safety,weather, Modules[pv],Deck[pv],x1,y1,z1,ll,id,xx,yy,zz,di,di/60,PFD))
            # print("{:50s} {:5s} {:4s} {:10s} {:8.1f} {:8.1f} {:8.1f} ".format(scn, Modules[pv],Deck[pv],id,di,di/60,PFD))            
            DimScn[(id,jet_or_pool)] = [pv,hole,fgdet,safety,weather, Modules[pv],Deck[pv],x1,y1,z1,ll,xx,yy,zz,di,di/60,PFD,np.nan,np.nan]
            if cube_result2file == True:
                print("{:50s} {:5s} {:4s} {:10s} {:8.1f} {:8.1f} {:8.1f} ".\
                    format(scn, Modules[pv],Deck[pv],id,di,di/60,PFD),file=f_cube_result)                
                print_cum_cube_file(id,IDAsorted,f_cube_result)  
                DimCubeSuccessList[id] = "'tvd-"+pv+"_"+hole             

        else: #Jet fire it is!
            pv,hole,weather = scn.split("\\")            
            weather_fire = weather            
            x1 = X[pv]
            y1 = Y[pv]
            z1 = Z[pv]
            z1 = z1+35
            for e in lEvent:
                if e.Key in scn:
                    break
            rrifound = False
            rr0 = 0.
            rri = 0.
            rr5min = 0.
            for i in range(0,len(e.TVD)-1):
                if (e.TVD[i,0] < di) and (e.TVD[i+1,0] >= di):
                    rr0 = e.Discharge.RRs[0]
                    rri = e.TVD[i,2]
                    rr5min = e.Discharge.RRs[1]
                    rrifound = True
                    break
            if rrifound == False:
                print(id,"Something wrong, rri not found", di, e.TVD[-1,0])

            dx=xx-x1
            dy=yy-y1
            dz=zz-z1
            ll = math.sqrt(dx*dx+dy*dy+dz*dz)
            # jli = jffit(rri)
            # xm = x1 + 0.61*dx
            # ym = y1 + 0.61*dy
            # zm = z1 + 0.61*dz
            # print("IS: {:50s} {:4s}{:2s}{:6.1f}{:6.1f}{:6.1f}  Cube: {:10s}{:6.1f}{:6.1f}{:6.1f} Dx:{:6.1f}{:6.1f}{:6.1f} {:8.1f}{:8.1f}".\
            #     format(scn,Modules[pv],Deck[pv],x1,y1,z1,id,xx,yy,zz,xx-x1,yy-y1,zz-z1,di,rri))

            #print with release rate at t=0,t=at tht time of leaving the cube, 5 min after the release
            pv,hole,weather = scn.split("\\")
            safety = hole[3:]    
            hole = hole[:2]
            weather,jet_or_pool = weather.split("_")
            fgdet = weather[-2:]
            weather = weather[:-2]
            if di > 300:
                print("{:10s} {:20s} {:2s} {:2s} {:4s} {:6s} {:5s} {:4s} ({:6.1f},{:6.1f},{:6.1f} ) -> {:6.1f} -> {:10s} ({:6.1f},{:6.1f},{:6.1f} ) {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f}".\
                    format(jet_or_pool,pv,hole,fgdet,safety,weather, Modules[pv],Deck[pv],x1,y1,z1,ll,id,xx,yy,zz, di,di/60,rr0,rr5min,rri))
                DimScn[(id,jet_or_pool)] = [pv,hole,fgdet,safety,weather, Modules[pv],Deck[pv],x1,y1,z1,ll,xx,yy,zz, di,di/60,rr0,rr5min,rri]
            else:
                print("{:10s} {:20s} {:2s} {:2s} {:4s} {:6s} {:5s} {:4s} ({:6.1f},{:6.1f},{:6.1f} ) -> {:6.1f} -> {:10s} ({:6.1f},{:6.1f},{:6.1f} ) {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f}".\
                    format(jet_or_pool,pv,hole,fgdet,safety,weather, Modules[pv],Deck[pv],x1,y1,z1,ll,id,xx,yy,zz, di,di/60,rr0,rri,np.nan))
                DimScn[(id,jet_or_pool)] = [pv,hole,fgdet,safety,weather, Modules[pv],Deck[pv],x1,y1,z1,ll,xx,yy,zz, di,di/60,rr0,rri,np.nan]

            DimCubeSuccessList[id] = "'tvd-"+pv+"_"+hole
            if cube_result2file == True:                
                print_cum_cube_file(id,IDAsorted,f_cube_result) 

            # print("IS: {:50s} {:4s}{:2s}{:6.1f}{:6.1f}{:6.1f}  Cube: {:10s}{:6.1f}{:6.1f}{:6.1f} Dx:{:6.1f}{:6.1f}{:6.1f} {:4s}{:2s} -> {:8s}{:8.1f}{:8.1f}{:8.1f}".\
            #     format(scn,Modules[pv],Deck[pv],x1,y1,z1,id,xx,yy,zz,xx-x1,yy-y1,zz-z1,Modules[pv],Deck[pv],id,di,rri,2.8893*np.power(55.5*rri,0.3728)))
            # print("IS: {:50s} {:6.1f}{:6.1f}{:6.1f}  Cube: {:10s}{:6.1f}{:6.1f}{:6.1f} Release rate: {:8.1f}".\
                # format(scn,x1,y1,z1,id,xx,yy,zz,rri))
            # print("SCYL: {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:6.1f} {:6.1f}".format(1000*x1,1000*y1,1000*z1,1000*xm,1000*ym,1000*zm,0,0.12*jli*1000))
            # print("SCYL: {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:6.1f} {:6.1f}".format(1000*xm,1000*ym,1000*zm,1000*xx,1000*yy,1000*zz,0.12*jli*1000,0))
            # print_cum_cube(id,IDAsorted) 
    elif not InterpolationSuccess:
        
        scn = '                                                                     No dimensioning scenario for ' + id
        print(scn)
        # print_cum_cube(id+" No dimensioning scenario ",IDAsorted)        
        if cube_result2file == True:
            print_cum_cube_file(scn,IDAsorted,f_cube_result)        


    # if (InterpolationSuccess == True) and (WantPlot == True):
    if (WantPlot == True):
        ll = len(IDAsorted)
        ec = np.zeros([ll,2])
        i=0
        ec[i,1] = IDAsorted[-1][0] #the longest duration
        ec[i,0] = IDAsorted[-1][1] #Frequency for the longest duration
        for i in range(1,len(IDAsorted)):
            ec[i,1] = IDAsorted[ll-i-1][0]
            ec[i,0] = ec[i-1,0] + IDAsorted[ll-i-1][1]

        # plt.figure(figsize=(5.91, 3.15))
        CF = ec[1:,0]
        JFL = ec[1:,1]
        masscolor = 'tab:blue'
        fig,ax1 = plt.subplots()
        ax1.set_xlabel('Jet Impingement Duration [sec]')
        ax1.set_ylabel('Cumulative Frequency [#/year]',color=masscolor)
        ax1.semilogy(JFL,CF,color=masscolor)
        ax1.set_ylim(top=5E-4,bottom=1E-6)
        ax1.set_xlim(left=0, right=3600)
        ax1.tick_params(axis='y',labelcolor=masscolor)
        ax1.xaxis.set_major_locator(plt.FixedLocator([300, 600, 1800, 3600]))
        ax1.annotate(scn,xy=(di,1.0E-4),xytext=(di,2E-4),horizontalalignment='left',verticalalignment='top',arrowprops = dict(facecolor='black',headwidth=4,width=2,headlength=4))
        # ax.xaxis.set_major_formatter(plt.FixedFormatter(['2/3','3/4','4/5','S05']))
        # ax.yaxis.set_major_locator(plt.FixedLocator([-27, -3.1, 3.1, 27]))
        # ax.yaxis.set_major_formatter(plt.FixedFormatter(['ER_S','Tray_S','Tray_P','ER_P']))
        ax1.grid(True,which="major")

        if InterpolationSuccess:
            plt.title("Cube {:10s} - {:5s} fire from {:30} for {:8.1f} [sec]".format(id,weather_fire[6:],pv+"_"+hole+"_"+weather_fire[:5],di))            
        else:
            plt.title("Cube {:10s} - No dimensioning fire".format(id))            
        plt.tight_layout()
        plt.show()

        fig.savefig("{}_{}.png".format(id,Version))
        plt.close()

    if cube_result2file == True:
        f_cube_result.close()

element_dump_filename_DimCube5 = element_dump_filename + '_DimCube5'
with open(element_dump_filename_DimCube5,'wb') as element_dump:
    dill.dump(lEvent,element_dump)

# print_cum_cube(IDAsorted)

# for c in Cubes:
#     print("{:20s} {:8.1f} [min]".format(c,CubeImpingeDuration[c]/60))
# sys.stdout.close()

# C = 0.
# for e in lEvent:
#     # if True:
#     # if e.Discharge.ReleaseRate*(1-e.Discharge.LiquidMassFraction) > 105:        
#     if e.Discharge.ReleaseRate > 105:
#         C += e.Frequency
# print("Leak Frequency at", Area, ":", C)
#  """
# """ 
# ModuleToCheck = "Hull"
# if Area == "HullDeck":
#     for a_or_f in ['A','F']:
#         CF = 0.
#         CFJ = 0.
#         CFEP = 0.
#         CFLP = 0.
#         print(a_or_f)
#         for e in lEvent:
#             if (e.Module == ModuleToCheck) and (((a_or_f == "A") and (e.X < X_hd_coaming))  or ((a_or_f == "F") and (e.X >= X_hd_coaming)) ):
#                 CFJ += e.JetFire.Frequency*6
#                 if e.EarlyPoolFire != None:
#                     CFEP += e.EarlyPoolFire.Frequency
#                 if e.LatePoolFire != None:
#                     CFLP += e.LatePoolFire.Frequency
#         print ("Fire accident frequency - Jet:", CFJ)
#         print ("Fire accident frequency - Early Pool:", CFEP)
#         print ("Fire accident frequency - Late Pool", CFLP)
#         print ("Fire accident frequency - Total", CFJ + CFEP + CFLP)



pdDS = pd.DataFrame(DimScn)
pdDS = pdDS.T
pdDS.columns = ['IS','Hole','F&G','ESD_BDV','Weather','Module','Deck','Xs','Ys','Zs','Distance','Xc','Yc','Zc','Sec','Min','PD_or_RR0','RR1','RR2']
pdDS.to_excel('Bv08_c5_DimScn_20201216.xlsx',float_format="%.1f",header=True,index=True)


listEPF = []
for e in lEvent:
    if (e.EarlyPoolFire != None):        
        key,hole,weather=e.Key.split("\\")
        Ds = e.EarlyPoolFire.PoolFireDurations
        listEPF.append([key, hole, weather,e.Module,e.Deck, e.EarlyPoolFire.Frequency,e.EarlyPoolFire.Diameter,Ds[0],Ds[1],Ds[2] ])
pdEPF = pd.DataFrame(listEPF,columns=['Key','Hole','Weather','Module','Deck','Frequency','Diameter','TimePoolRelease','TimePoolBurn','TimePoolFire'])
is_Ts = pdEPF.TimePoolFire > 0
pdEPF[is_Ts]

is_S02 = pdEPF.Module == 'S02'
is_FullCoaming = pdEPF.Diameter > 15
pdEPF[is_S02 & is_FullCoaming]



s6 = pd.read_csv('S05_A_AS_6.txt',sep="\s+",header = None,names=['Key','Module,','F','T','Fcum'])
s5 = pd.read_csv('S05_A_AS_5.txt',sep="\s+",header = None,names=['Key','Module,','F','T','Fcum'])
plt.semilogy(s5['T'],s5.Fcum,s6['T'],s6.Fcum)
plt.legend(['Pool or Jet','Pool/Jet Distributed'])
plt.show()