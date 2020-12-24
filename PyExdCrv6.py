#PyExdCrv6.py with D37 for Jet fire
#Ignition probability is updated considering 5% addition to the two=phase release also.
#PyExdCrv5.py with the updated frequency at S01

import matplotlib.pyplot as plt
from openpyxl import load_workbook
import math
import numpy as np
import dill
import unicodedata
import re
import sys


def print_a_event(lEvent,key='',hole='',weather=''):
    i=0
    for e in lEvent:
        if (key in e.Key) and (hole in e.Hole) and (weather in e.Weather):
            print(i,e.Key)
        i += 1

def slugify(value):
    """
    Normalizes string, converts to lowercase, removes non-alpha characters,
    and converts spaces to hyphens.
    """
    # import unicodedata
    # value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore')
    # value = unicode(re.sub('[^\w\s-]', '', value).strip().lower())
    # value = unicode(re.sub('[-\s]+', '-', value))
    value = re.sub('[^\w\s-]', '', value).strip().lower()
    value = re.sub('[-\s]+', '-', value)
    return value

NumDirections = 6 # 6 cones to cover the all release direction for a sphere

TVDPLot = False

# iExlFilename='Bv08_i5'
#Topside
# Area = 'ProcessArea'
# cExlFilename='Bv08_c5'
# EqvsModule = {'020-01': 'S05', '020-02':'S05', \
#     '020-03':'S04', '020-04':'S04', \
#         '021-01':'S02','023-01':'P03','023-02':'P03', \
#             '023-03':'S03', '023-04':'S03', \
#                 '024-01':'P04', '024-02':'P04','024-03':'P04', '025-01':'P04', '025-02':'P04', \
#                     '025-03':'S03','027-01':'P05', '027-02':'P05', '027-03':'P05', \
#                         '045-01':'P03','045-02':'P03','046-01':'P02','013-01':'S05'}

# SFXFiles = ['013-01-C1-G', '013-01-C2-G', '013-01-N1-G', '013-01-N2-G', \
#     '020-01-01-G', '020-01-02-L', '020-02-01-Li', '020-02-02-G', '020-02-03-Lo', '020-03-01-Li', '020-03-02-G', '020-03-03-Lo', '020-04-01-L', '020-04-02-G', '020-05-01-G', '020-05-02-L', \
#     '023-01-01-G-DA', '023-01-01-G-DB', '023-01-02-L', '023-02-01-G-DA', '023-02-01-G-DB', '023-02-02-L', '023-03-01-G-DA', '023-03-01-G-DB', '023-03-01-G-DC', '023-03-02-L','023-03-03-L', '023-04-G',\
#     '024-01-01-G','024-01-02-L', '024-02-G-DA', '024-02-G-DB', '024-02-G-DC', '024-03-01-G', '024-03-02-L',  \
#     '025-01-01-G', '025-01-02-L', '025-02-01-G', '025-02-02-L', '027-01-G-DA', '027-01-G-DB', '027-01-G-DC', '027-02-G', '027-03-G',  \
#     '043-03-G', '045-01-G', '045-02-01-G', '045-02-02-L', '045-03-G', '046-02-L', '046-03-L',\
#     '021-01-01-L','046-01-01-L','046-01-02-L','046-01-03-L','046-01-04-L','062-01-01-L','043-01-01-G','043-01-02-L','043-02-01-G','043-02-02-L']
# TVDprefix = '.\\tvd_rev.B\\tvd-xlsx & pngs\\'
# element_dump_filename = 'Bv08_c5_dump'


#Topside Parameters End

#Offloaidng area analysis
# 'Flammable dispersion' is not relevant??? All vaporized??
Area = 'Offloading'
cExlFilename='Bv06_O_c'
SFXFiles = ['021-02-L']
element_dump_filename = 'Bv06_offloading_dump'

#Utlity analysis
# Area = 'Utility'
# cExlFilename='Bv06_u_c'
# SFXFiles = ['045-04-G','062-01-02-L']
# TVDprefix = '.\\tvd_rev.B\\tvd-xlsx & pngs\\'
# element_dump_filename = 'Bv06_utility_dump_2' #_2 with the updated S01 part count
# element_dump_filename = 'Bv06_utility_dump' #_2 with the updated S01 part count

#Hull deck analysis
# iExlFilename='Bv06_i'
# Area = 'Hull'
# cExlFilename='Bv06_h_c'
# SFXFiles = ['131-01-L-fwdP','131-01-L-fwdS','131-01-L-aftP','131-01-L-aftS','131-02-L-fwdP','131-02-L-fwdS','131-02-L-aftP','131-02-L-aftS','057-01-02-L','057-01-01-G','046-04-L','021-01-02-L','013-06-L','013-05-L']
# TVDprefix = '.\\tvd_rev.B\\tvd-xlsx & pngs\\'
# element_dump_filename = 'Bv06_hull_dump'



iExl=load_workbook(filename=iExlFilename+'.xlsx')
cExl=load_workbook(filename=cExlFilename+'.xlsx')
fExl=load_workbook(filename='Leak_Freq_Hole_Size.xlsx')
# fExl=load_workbook(filename='Leak_Freq_Hole_Size_2.xlsx')
shFreq = fExl['RESULTS - Hole size']


shPV = iExl['Pressure vessel']
shBund = iExl['Type of pool substrate and bund']
shtTLV = iExl['Time varying leak']
shDc = cExl['Discharge']
shJet = cExl['Jet fire']
shDispersion = cExl['Flammable Dispersion']
shExplosion = cExl['Explosions']


def jffit(m):
    jl_lowe = 2.8893*np.power(55.5*m,0.3728)
    # if m>5:
    #     jf = -13.2+54.3*math.log10(m)
    # elif m>0.1:
    #     jf= 3.736*m + 6.
    # else:
    #     jf = 0.
    # print(m, jl_lowe,jf)
    return jl_lowe
def mfit(jl):
    m = np.power(10,math.log10(jl/2.8893)/0.3728) / 55.5
    
    return m

class Event:
    def __init__(self,study_folder_pv,scenario,weather,dc=None,jf=None,epf=None,lpf=None,exp=None,freq=None,x=None,y=None,rh = None, \
        hole = None,holemm = None, PP = None, TT = None,fireball = None, dispersion = None, esd = True, bdv = True,module = None,deck= None,pesd = 1, pbdv = 1):        
        self.Study_Folder_Pv = study_folder_pv
        self.Path = study_folder_pv + "\\" + scenario
        self.Weather = weather
        study,pv, folder = study_folder_pv.split("\\")
        self.Key = pv + "\\" + scenario + "\\" + weather
        self.Frequency = freq
        self.Discharge = dc
        self.JetFire = jf
        self.EarlyPoolFire = epf
        self.LatePoolFire = lpf
        self.Explosion = exp
        self.Fireball = fireball
        self.Dispersion = dispersion
        self.TVD = []
        self.TVDRead = False
        self.Module = module
        self.Deck = deck
        self.X = x
        self.Y = y
        self.ReleaseHeight = rh #should be 'Elevation.' Or Elevation + ReleaseHeight?
        self.Hole = hole
        self.Holemm = holemm
        self.Pressure = PP
        self.Temperature = TT        
        self.PImdIgn = 1
        self.PDelIgn = 0
        self.PExp_Ign = 0.21 #WOAD as referred by by IP & UKOOA as the 'probability of explosion given ignition'
        self.ESDSuccess = esd
        self.BDVSuccess = bdv
        self.BDVDia = 0
        self.jfscale = 1      
        self.PESD = pesd
        self.PBDV = pbdv
        self.BundArea = 0.
        """     def __str__(self):
        ts, folder, pv, scenario = self.Path.split("\\")
        if self.TVDRead == True:
            # RelDuration = self.TVD[-1,0]
            a,b,c,d,e = self.JetFire.JetLengths
            fmt = "{:10s} | {:40s} | {:20s} | {:6.2f} |{:6.2f} |{:6.2f} | {:6s} | {:8s} | {:6s} | {:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}*".\
                format(ts,   folder,      pv, self.X, self.Y, self.ReleaseHeight, self.Module, scenario, self.Weather[9:], self.JetFire.Frequency, self.JetFire.SEP, a, b, c, d,e)
        else:
            fmt = "{:10s} | {:40s} | {:20s} | {:20s} | {:20s} | {:20s} | {:.2e} | {:8.2f} | {:8.2f}".\
                format(ts,   folder,      pv, self.Module, scenario, self.Weather, self.Frequency, self.Discharge.ReleaseRate,self.Discharge.Duration)
        
        return fmt """
    def __str__(self):
        ts, folder, pv, scenario = self.Path.split("\\")
        if self.TVDRead == True:
            # print(self.Path,self.Dispersion,self.JetFire,self.EarlyPoolFire,self.LatePoolFire,self.Explosion,self.Fireball)
            # RelDuration = self.TVD[-1,0]
            # a,b,c,d,e = self.JetFire.JetLengths
            fmt = "{:20s} | {:6.2f} |{:6.2f} |{:6.2f} | {:6s} | {:8s} | {:6s} \
                | {:8.2f} [barg]| {:8.2f} [degC]| | ESD : {:}| BDV  {:} | Freq: {:8.2e}\n".\
                format(pv, self.X, self.Y, self.ReleaseHeight, self.Module, scenario, self.Weather[9:], \
                self.Pressure, self.Temperature, self.ESDSuccess,self.BDVSuccess, self.Frequency)
            
            if self.Dispersion != None:
                fmt += "\tDispersion: Frequency {:8.2e}, Distance to LFL {:}\n".format(self.Dispersion.Frequency,self.Dispersion.DfMax)
            else:
                fmt += "\tDispersion: No dispersion\n"
            
            if self.JetFire != None:
                fmt += "\tJF: Frequency {:8.2e}, Length {:6.1f}\n".format(self.JetFire.Frequency,self.JetFire.Length)
            else:
                fmt += "\tJF: No Jet Fire\n"
            
            if self.Explosion != None:
                fmt += "\tExplosion: Frequency {:8.2e}, Overpressure {:6.1f} to {:6.1f} [m]\n".format(self.Explosion.Frequency,self.Explosion.Overpressures[-1],self.Explosion.Distance2OPs[-1])
            else:
                fmt += "\tExplosion: No Explosion\n"
            
            if self.EarlyPoolFire != None:
                fmt += "\tEPF: Frequency {:8.2e}, Diameter {:6.1f}\n".format(self.EarlyPoolFire.Frequency,self.EarlyPoolFire.Diameter)
            else:
                fmt += "\tEPF: No Early Pool Fire\n"
            if self.LatePoolFire != None:
                fmt += "\tLPF: Frequency {:8.2e}, Diameter {:6.1f}\n".format(self.LatePoolFire.Frequency,self.LatePoolFire.Diameter)
            else:
                fmt += "\tLPF: No Late Pool Fire\n"
            if self.Fireball != None:
                fmt += "\tFireball: Frequency {:8.2e}, Diameter {:6.1f}\n".format(self.Fireball.Frequency,self.Fireball.Diameter)
            else:
                fmt += "\tFireball: No Fireball\n"
        
        else:
            ts,pv,IS,hole = self.Path.split("\\")
            f = "{:15s} {:4s} ({:6.1f},{:6.1f},{:6.1f}) {:6.1f} {:6s} P({:8.2f}) T({:8.2f}) ESD({:}) BDV({:}) | Freq: {:8.2e}".\
                        format(IS, self.Module, self.X, self.Y,  self.ReleaseHeight, self.Holemm, self.Weather,self.Pressure, self.Temperature, self.ESDSuccess,self.BDVSuccess,self.Frequency)
            if self.Dispersion != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.JetFire != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.EarlyPoolFire != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.LatePoolFire != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.Explosion != None:
                f += "  O  "
            else:
                f += "  X  "
            if self.Fireball != None:
                f += "  O  "
            else:
                f += "  X  "            
            fmt = f
        
        return fmt
        # return self.Key
            

class Discharge:
    def __init__(self,rr,duration,liqfrac = 0.):
        self.ReleaseRate = rr
        self.Duration = duration
        self.LiquidMassFraction = liqfrac
        self.Ts = [0.,0.,0.,0.,0.]
        self.Ms = [0.,0.,0.,0.,0.]
        self.RRs = [0.,0.,0.,0.,0.]        

class Dispersion:
    def __init__(self,fdisp,dfmax,dfmin,wf,dmax,dmin,w,ffmax,ffw):
        self.Frequency = fdisp        
        
        # self.Distance2LFLFraction = d2f
        # self.Distance2LFL = d
        
        #Column P: max d, Q: min d, R: max w for LFL fraction
        #Column S: max d, T: min d, U: max w for LFL fraction
        self.DfMax = dfmax
        self.DfMin = dfmin
        self.Wf = wf
        self.DMax = dmax
        self.DMin = dmin
        self.W = w

        self.FFMaxDistance = ffmax
        self.FFWidth = ffw

    def __str__(self):
        
        fmt = "{:.2e} | {:8.2f}m to LFL Fraction | {:8.2f}m to LFL".\
            format(self.Frequency, self.DfMax, self.DMax)        
        return fmt

class JetFire:
    def __init__(self,length,sep,jff,d04,d12,d37):
        self.Length = length
        self.SEP = sep
        self.Frequency = jff
        self.D04 = d04
        self.D12 = d12
        self.D37 = d37
        self.Ts = [0.,0.,0.,0.,0.]
        self.JetLengths = [0.,0.,0.,0.,0.]
    def __str__(self):
        a,b,c,d,e = self.JetLengths
        fmt = "{:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}".\
            format(self.Frequency, self.SEP, a, b, c, d, e)        
        return fmt

class EarlyPoolFire:
    def __init__(self,diameter,sep,epff,d04,d12):
        self.Diameter = diameter
        self.SEP = sep
        self.Frequency = epff
        self.D04 = d04 #Distance to 4kW/m2
        self.D12 = d12 #Distance to 12.4 kW/m2
        self.Ts = [0.,0.,0.,0.,0.]
        self.PoolFireDurations = [0.,0.,0.]
        self.PoolDiameters = [0.,0.,0.,0.,0.]
    def __str__(self):
        a,b,c,d,e = self.PoolDiameters
        fmt = "{:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}".\
            format(self.Frequency, self.SEP, a, b, c, d, e)

class LatePoolFire:
    def __init__(self,diameter,sep,epff,d04,d12):
        self.Diameter = diameter
        self.SEP = sep
        self.Frequency = epff
        self.D04 = d04 #Distance to 4kW/m2
        self.D12 = d12 #Distance to 12.4 kW/m2
        self.Ts = [0.,0.,0.,0.,0.]
        self.PoolDiameters = [0.,0.,0.,0.,0.]
    def __str__(self):
        a,b,c,d,e = self.PoolDiameters
        fmt = "{:.2e} | {:6.0f} | {:8.2f} | {:8.2f} | {:8.2f} | {:8.2f}| {:8.2f}".\
            format(self.Frequency, self.SEP, a, b, c, d, e)

class Explosion:
    def __init__(self,fexp,op,d2op):
        self.Overpressures = op
        self.Distance2OPs = d2op
        self.Frequency = fexp        
    def __str__(self):
        a,b,c = self.Distance2OPs
        pa,pb,pc = self.Overpressures
        fmt = "{:.2e} | {:8.2f}barg to {:8.2f} | {:8.2f}bar to {:8.2f} | {:8.2f} barg to {:8.2f}".\
            format(self.Frequency, pa, a, pb, b, pc,c)        
        return fmt

class Fireball:
    def __init__(self,ffb,D,sep):
        self.Diameter = D
        self.SEP = sep
        self.Frequency = ffb       
    def __str__(self):
        
        fmt = "{:.2e} | Diameter: {:8.2f} | SEP {:8.2f}".\
            format(self.Frequency, self.Diameter, self.SEP)        
        return fmt

#Read Leaks
#listLeak = []
ReleaseHeight= {} #from vessel bottom, Column 'V'? Shall be elevation Columnt 'T'?
Frequency ={}
PImdIgn ={}
PDelIgn = {}
PExp ={}
ReleaseRate = {}
Duration = {}
J00 = {}
J05 = {}
J10 = {}
J30 = {}
J60 = {}
JetSEP = {}
JetFrequency = {}
Tiso = {}
Tbdv = {}
Hole = {}
Holemm = {}
PP = {} #Pressure of the PV
TT  = {} #Temp of the PV
ESDSuccess ={}
BDVSuccess ={}
BDVDia ={}
X = {}
Y = {}
Z = {}
Material = {}
Bund = {}



# iIS=load_workbook(filename='IS_v12_shk.xlsx',data_only=True)
iIS=load_workbook(filename='IS_v15_for_SHI_20201127.xlsx',data_only=True)
shIS = iIS['Isolatable summary']
IS_sub = {}
numESDVs = {}
Modules = {}
Deck = {}
IS_P = {}
IS_T = {}
IS_rho = {}
IS_V = {}
IS_Vadj = {}
r = 3
# while r < 79:
#     nsub = shIS.cell(r,4).value
#     IS_sub[shIS.cell(r,3).value] = [r,nsub]
#     r += nsub
while r < 81:
    #for IS_v12
    # pv = shIS.cell(r,5).value
    # IS_sub[pv] = shIS.cell(r,11).value #Read for each leak at respective height
    # Modules[pv] = shIS.cell(r,7).value
    # Deck[pv] =  shIS.cell(r,8).value

    # IS_P[pv] = shIS.cell(r,38).value
    # IS_T[pv] = shIS.cell(r,39).value
    # IS_rho[pv] = shIS.cell(r,40).value
    # IS_V[pv] = shIS.cell(r,32).value
    # IS_Vadj[pv] = shIS.cell(r,33).value    
    # if shIS.cell(r,24).value != None:
    #    nedvs = shIS.cell(r,24).value.count("\"")    
    #    numESDVs[pv] = nedvs
    # else:
    #    numESDVs[pv] = pnedvs
    #For IS_v15
    pv = shIS.cell(r,3).value
    IS_sub[pv] = shIS.cell(r,9).value #Read for each leak at respective height
    if IS_sub[pv] == 'NA':
        IS_sub[pv] = 1         
    Modules[pv] = shIS.cell(r,5).value
    Deck[pv] =  shIS.cell(r,6).value

    IS_P[pv] = shIS.cell(r,39).value
    IS_T[pv] = shIS.cell(r,40).value
    IS_rho[pv] = shIS.cell(r,41).value
    IS_V[pv] = shIS.cell(r,33).value
    IS_Vadj[pv] = shIS.cell(r,34).value    
    if shIS.cell(r,23).value != None:
       nedvs = shIS.cell(r,23).value.count("\"")    
       numESDVs[pv] = nedvs
    else:
       numESDVs[pv] = pnedvs
    pnedvs = nedvs
    r += 1

import pandas as pd
pdIS = pd.DataFrame([Modules, Deck,IS_P, IS_T,IS_V,IS_Vadj],index=['Modules','Deck','P','T','Vol','VolAdj'])
pdIS = pdIS.transpose()
pdIS = pdIS.replace(0,'NaN')
pdIS = pdIS.fillna(method='ffill')
pdISM = pdIS.set_index(['Modules','Deck'])
#Pressure analysis
# pdISM.mean(level=['Modules','Deck']).P.plot.bar() #Average pressure per Module-Deck
# plt.title('Pressure')
# plt.show()
#Volume analysis
pdISM.eval('Vtotal = Vol + VolAdj')
# pdISM.mean(level=['Modules','Deck']).P.plot.bar() #Average pressure per Module-Deck
# plt.show()


r=5
while shFreq.cell(r,2).value == "Full pressure leak":     
    pv  = shFreq.cell(r,1).value 
    Frequency[pv] = [0.,0.,0.,0.]
    if pv[-1] == "G":
        Frequency[pv][0] = shFreq.cell(r,4).value #Gas SM
        Frequency[pv][1] = shFreq.cell(r,5).value # ME
        Frequency[pv][2] = shFreq.cell(r,6).value # MA
        Frequency[pv][3] = shFreq.cell(r,7).value # LA
    else:
        Frequency[pv][0] = shFreq.cell(r,10).value #Liq SM
        Frequency[pv][1] = shFreq.cell(r,11).value # ME
        Frequency[pv][2] = shFreq.cell(r,12).value # MA
        Frequency[pv][3] = shFreq.cell(r,13).value # LA
    r += 3
numIS = (r-5)/3


# shBund = iExl['Type of pool substrate and bund']

#Read 'Presure vessel' from Input
r=63
while shPV.cell(r,1).value == "Yes":    
    study  = shPV.cell(r,2).value
    pv  = shPV.cell(r,8).value    
    # key = study  + "\\" +  pv
    key = pv
    Material[key] = shPV.cell(r,9).value
    TT[key] = shPV.cell(r,15).value
    PP[key] = shPV.cell(r,16).value
    X[key] = shPV.cell(r,136).value
    Y[key] = shPV.cell(r,137).value    
    Z[key] = shPV.cell(r,20).value        #Elevation
    if shPV.cell(r,62).value != None:
        BundType,BundName = shPV.cell(r,62).value.split('\\')
        for rb in range(63,shBund.max_row+1):
            if BundName == shBund.cell(rb,3).value:
                Bund[key] = shBund.cell(rb,6).value
                break
    r += 1
numPV = r-63




#Read 'Time varying leak' from Input
#for r in range(6,numscn):
r = 63
while shtTLV.cell(r,1).value == "Yes":
    study = shtTLV.cell(r,2).value    
    relheight = shtTLV.cell(r,32).value
    equip  = shtTLV.cell(r,8).value
    folder = shtTLV.cell(r,9).value
    hole = shtTLV.cell(r,20).value 
    holemm = shtTLV.cell(r,21).value
    relheight = shtTLV.cell(r,25).value
    # key = study + "\\" + equip + "\\" + folder + "\\"  + hole
    key_hole = equip+"-"+hole
    ReleaseHeight[equip] = relheight

    #Construci disctionarys with keys as 'Path'
    Hole[key_hole] = hole
    Holemm[key_hole] = holemm
    #SAFETI
    # Frequency[key]  = shtTLV.cell(r,38).value
    # PImdIgn[key] = shtTLV.cell(r,44).value
    # PDelIgn[key] = shtTLV.cell(r,46).value
    # PExp[key] = shtTLV.cell(r,48).value
    #PHAST doesn't provide this frequency & probability    
    
    # PHAST
    ESDSuccess[key_hole] = shtTLV.cell(r,44).value
    Tiso[key_hole] = shtTLV.cell(r,45).value
    BDVSuccess[key_hole] = shtTLV.cell(r,47).value
    Tbdv[key_hole] = shtTLV.cell(r,48).value
    BDVDia[key_hole] = shtTLV.cell(r,49).value
    r=r+1
    print("Reading Time varying leaks: ",key_hole)
numLeak = r-63



lEvent = [] #List of class Event
#Construct Event list
#Read Discharge, shDc
# And evaluate ignition probability'
# weathers = ['1.2/F', '6.6/D']
weathers = ['2.9F','7.7D','14.5D']
HoleSizes = [7.1, 36.1, 111.8, 150]
dk=0
#for r in range(2,numLeak*len(weathers)+2):
for r in range(2,shDc.max_row+1):
    study_folder_pv = shDc.cell(r,1).value
    hole = shDc.cell(r,2).value
    if hole != "Catastrophic rupture":
        study,pv,folder = study_folder_pv.split("\\")
        scenario = shDc.cell(r,2).value
        weather = shDc.cell(r,3).value
        path = study_folder_pv + "\\" + scenario
        # key = study_folder_pv + "\\" + scenario + "\\" + weather
        key_hole = pv +"-"+ hole
        key_hole_weather = pv +"-"+hole+"-"+weather
        liqmassfrac = shDc.cell(r,14).value
        if weather in weathers:
            rr = shDc.cell(r,10).value
            ReleaseRate[key_hole] = rr
            duration = shDc.cell(r,11).value
            x = X[pv]
            y = Y[pv]
            z = Z[pv]
            Duration[key_hole] = duration
            basefreq = 0.
            #Frequency allocatin
            if hole[0:2] in ['SM','ME','MA','LA']:
                holesizeindex = ['SM','ME','MA','LA'].index(hole[0:2])
                if pv in Frequency.keys():                
                    basefreq = Frequency[pv][holesizeindex]
                    holepath = 11
                else:
                    for k in Frequency.keys():
                        if k in pv:
                            fkey = k
                    basefreq = Frequency[fkey][holesizeindex]
                    holepath = 12
                    
            elif hole[0:2] == "LM":
                for k in range(0,4):
                    if Holemm[key_hole] > HoleSizes[k]:
                        continue
                    else:
                        break
                holesizeindex = k                
                if pv in Frequency.keys():
                    for k in range(holesizeindex,4): #Add all frequency for larger hole size categories for LM
                        basefreq += Frequency[pv][k]
                        Frequency[pv][k] = 0.
                    Frequency[pv][holesizeindex] = basefreq                    
                    holepath = 21
                else:
                    for k in Frequency.keys():
                        if k in pv:
                            fkey = k
                    for k in range(holesizeindex,4):
                        basefreq += Frequency[fkey][k]
                        Frequency[fkey][k] = 0.
                    Frequency[fkey][holesizeindex] = basefreq
                    holepath = 22
            else:
                print(pv, hole[:2],'Something wrong in reading frequency')
                sys.exit()
            
            if '024-02-G' in pv:
                print(pv, hole[0:2],holesizeindex,basefreq,IS_sub[pv],holepath)

            leak_freq = IS_sub[pv]*basefreq#actually Frequency[pv] = 1            

            
            # aEvent = Event(study_folder_pv,scenario,weather,freq=Frequency[path],dc=Discharge(rr,duration),hole=Hole[path])
            p_ESDfail = 0.01*numESDVs[pv]
            if hole[4]=="O":
                esd = True
                p_esd_branch = 1-p_ESDfail
                leak_freq *= 1-p_ESDfail
            elif hole[4] == "X":
                esd = False
                p_esd_branch = p_ESDfail
                leak_freq *= p_ESDfail
            else:
                print(pv,hole,'ESDV value wrong')
                break
            #BDV failure into Event Frequency
            if hole[-1]=="O":
                bdv = True
                p_bdv_branch = (1 - 0.005)
                leak_freq *= (1 - 0.005)
            elif hole[-1]=="N":
                bdv = "None"
                p_bdv_branch = 1
            elif hole[-1]=="X":
                bdv = False
                p_bdv_branch = 0.005
                leak_freq *= 0.005
            else:
                print(pv,hole,'BDV value wrong')
                break

            if weather == '2.9F':
                leak_freq *= 0.486
            elif weather == '7.7D':
                leak_freq *= 0.471
            elif weather == '14.5D':
                leak_freq *= 0.044
            else:
                print('Wrong wetaher', key_hole_weather)
                break



            aEvent = Event(study_folder_pv,hole,weather,freq=leak_freq,dc=Discharge(rr,duration,liqfrac=liqmassfrac),hole=Hole[key_hole],holemm = Holemm[key_hole],\
                esd=esd,bdv=bdv,x=x,y=y,rh=z,PP=PP[pv],TT=TT[pv],module=Modules[pv],deck=Deck[pv],pesd=p_esd_branch, pbdv=p_bdv_branch)
            lEvent.append(aEvent)

        else:
            print("{} Discharge events read".format(dk))
            break
        dk=dk+1

#End of constructing list of Events
#Up to now, read IS, '_i', '_c[Discharge]'

IgnProb = {}
#Rev. A ERA
# IgnProb['023-01']=[0.00050007132890309,0.00120241794400881,0.0121099047419176,0.0133941985161508]
# IgnProb['023-02']=[0.000500009730687852,0.00112728346839808,0.0114299096696923,0.0124894860472479]
# IgnProb['045-01']=[0.000500092602037187,0.00214457708972392,0.0159824850536888,0.015707932976315]
# IgnProb['024-01']=[0.000500049238017383,0.00135755735655058,0.0127160815903571,0.0137890702655233]
# IgnProb['024-02']=[0.000500092602037187,0.00177918758871906,0.014582259955286,0.0152682624925632]
# IgnProb['024-03']=[0.000500092602037187,0.00187273284139629,0.0148864999002958,0.0153843297225681]
# IgnProb['025-01']=[0.000500066037325877,0.00155843931004843,0.0138293983168883,0.0148186113336855]
# IgnProb['025-02']=[0.000500076839107057,0.001583080559352,0.0138624647673917,0.0148213944403002]
# IgnProb['027-01']=[0.000500092602037187,0.00141234355179696,0.0130284920753651,0.014098525976969]
# IgnProb['027-02']=[0.000500092602037187,0.00167828753177232,0.0139540201469402,0.0148124734230255]
# IgnProb['023-03']=[0.000500074056674887,0.00134865578738282,0.0127446106070196,0.013756233693594]
# IgnProb['023-04']=[0.000500014189107731,0.00114561328550342,0.0121656333278765,0.0136982452845986]
# IgnProb['024-01']=[0.000500056109924782,0.00140883072048731,0.013214286058242,0.0148872691410471]
# IgnProb['045-02']=[0.000500077778731805,0.00141477302560298,0.012839767974552,0.0139304583821647]
# IgnProb['020-01']=[0.000500081597970954,0.00198507390262245,0.0159033776172752,0.0159730079240214]
# IgnProb['020-02']=[0.000500044233038885,0.00127252532590165,0.0126086841965461,0.0139360824741458]
# IgnProb['013-01']=[0.000500081597970954,0.00125827417928573,0.0124609025752969,0.0136146062289823]
# IgnProb['020-03']=[0.000500037598733664,0.00117389329582008,0.0122180292562774,0.0135140045120202]
# IgnProb['020-04']=[0.000500009985715746,0.00113070577630583,0.0120388416453966,0.0131069382140561]
# IgnProb['027-03']=[0.000500087509063428,0.00137082995866607,0.0129208012390835,0.0142136300044506]
# IgnProb['General']=[0.000502,0.001677,0.017112,0.017916]

IgnProb['013-01-N1-G']=[0.000500741303319378, 0.00173837090794396, 0.0162488653124486, 0.0256593965924138]
IgnProb['013-01-N2-G']=[0.000500732618994306, 0.00173021026054981, 0.0162106686883042, 0.0233365963929737]
IgnProb['013-01-C1-G']=[0.000500730057951166, 0.00172839862201963, 0.0161280460062742, 0.0260396368845274]
IgnProb['013-01-C2-G']=[0.000500732618994306, 0.00173040063784021, 0.0162372735346198, 0.0231392045385233]
IgnProb['020-01-01-G']=[0.00050078082748584, 0.0022512591356413, 0.0184538666248445, 0.0349498349619589]
IgnProb['020-02-02-G']=[0.000500686531289899, 0.00206330653062102, 0.0156159389874572, 0.0193917750864195]
IgnProb['020-03-02-G']=[0.000500614151127518, 0.00162108974615963, 0.0143712153162215, 0.0164269326631637]
IgnProb['020-04-02-G']=[0.00050061223221902, 0.00155259472908994, 0.013872912082411, 0.0154985638645694]
IgnProb['020-05-01-G']=[0.000500701526556603, 0.00178692749944163, 0.0159702151117313, 0.0299781987342712]
IgnProb['023-01-01-G-D']=[0.00050064954574844, 0.00164554571797046, 0.0144585554126121, 0.0165481589146481]
IgnProb['023-02-01-G-D']=[0.000500722070003689, 0.00176489465227185, 0.0151442544227508, 0.0184920435195054]
IgnProb['023-03-01-G-D']=[0.000500721623901645, 0.00210758706763709, 0.018099791610662, 0.0233247729623398]
IgnProb['023-04-G']=[0.000500647466646291, 0.00167347526470548, 0.0150550133376746, 0.0182726178238411]
IgnProb['024-01-01-G']=[0.000500851793976426, 0.0021068526287551, 0.017749482850647, 0.0328857442445363]
IgnProb['024-02-G-D']=[0.000500777926955983, 0.00204218053256642, 0.0175090375222382, 0.0328080245173365]
IgnProb['024-03-01-G']=[0.000500777926955984, 0.00204894343672028, 0.0175472283187267, 0.0328130535364256]
IgnProb['025-01-01-G']=[0.000500699767165198, 0.00187828964183312, 0.0168039281164306, 0.0316443977210601]
IgnProb['025-02-01-G']=[0.000500800394970362, 0.00216928835315792, 0.0169698657622736, 0.0328017209524779]
IgnProb['027-01-G-D']=[0.000500765680124118, 0.00200820277437213, 0.0175694012489746, 0.0305038885269955]
IgnProb['027-02-G']=[0.000500672879178397, 0.00170571103988452, 0.015954339055376, 0.0301326015799154]
IgnProb['027-03-G']=[0.000500695169373991, 0.00174886765510705, 0.0161703485823794, 0.0311855056758914]
IgnProb['043-03-G']=[0.000500580883408993, 0.00120756122631826, 0, 0]
IgnProb['045-01-G']=[0.000500618600690852, 0.00159733916169166, 0.0153114338463327, 0.0296140266897886]
IgnProb['045-02-01-G']=[0.00050067402955842, 0.00204334635611344, 0.0163766033073665, 0.0202061053466699]
IgnProb['045-03-G']=[0.000500879020195334, 0.00140151702998937, 0.012180401004956, 0.0138600223818198]
IgnProb['043-01-01-G']=[0.000500602143837139, 0.00150199990443012, 0.013190074148517, 0.0147119415600005]
IgnProb['043-02-01-G']=[0.000500586565431687, 0.00133550843897021, 0.0116663326010593, 0.0136429022039804]
IgnProb['020-01-02-L']=[0.000522311709849998, 0.00101233411540825, 0.0100166030156997, 0.0150630910520578]
IgnProb['020-02-01-Li']=[0.000521670334120392, 0.00101236287935389, 0.0100154462606476, 0.0147669909395595]
IgnProb['020-02-03-Lo']=[0.000508794219862845, 0.00101244265441222, 0.0100185528257358, 0.0152672837144517]
IgnProb['020-03-01-Li']=[0.000530557960555035, 0.00101235268867158, 0.010015072804113, 0.0142935621996707]
IgnProb['020-03-03-Lo']=[0.000508700953213926, 0.00101246450879569, 0.0100204560901898, 0.0153256858965408]
IgnProb['020-04-01-L']=[0.00050870661406958, 0.00101249096486613, 0.0100230203211711, 0.0153413609392222]
IgnProb['020-05-02-L']=[0.000517639319296904, 0.00101232590565284, 0.0100168043711745, 0.0150339074995574]
IgnProb['023-01-02-L']=[0.00051765619051897, 0.00101236108289505, 0.0100138442263456, 0.0114191580370443]
IgnProb['023-02-02-L']=[0.000519364709638211, 0.00101247154623826, 0.0100138836975031, 0.0116845473593426]
IgnProb['023-03-02-L']=[0.000518883017438573, 0.00101233480725284, 0.0100138836975031, 0.0118524801475187]
IgnProb['023-03-03-L']=[0.00051612967735141, 0.0010122745346684, 0.0100138836975031, 0.0117661106021721]
IgnProb['024-01-02-L']=[0.000508737117440372, 0.00101232290320598, 0.0100138839172225, 0.0127603066014628]
IgnProb['024-03-02-L']=[0.000508786229423166, 0.00101235169916876, 0.0100138524038315, 0.0114198897932327]
IgnProb['025-01-02-L']=[0.000520000046982092, 0.00101239557824715, 0.0100138894846758, 0.0133825980638876]
IgnProb['025-02-02-L']=[0.000522583069562832, 0.00101236355066008, 0.0100138852135714, 0.0133282685238919]
IgnProb['045-02-02-L']=[0.000508796740718446, 0.00101235169916876, 0.0100138620131291, 0.0114198897932327]
IgnProb['046-02-L']=[0.000508730146956579, 0.00101235169916876, 0.0100138836975031, 0.0120016098996878]
IgnProb['046-03-L']=[0.000508860004113257, 0.00101235169916876, 0.0100138836975031, 0.0118524777688686]
IgnProb['021-01-01-L']=[0.000508858817833085, 0.00101259968232398, 0.0100138836975031, 0.011479454586573] #Condensate metering to be used for 'General'
IgnProb['046-01-01-L']=[0.000520054733740898, 0.0010124262251176, 0.0100138836975031, 0.0115909676594336]
IgnProb['046-01-02-L']=[0.00050866129882744, 0.00101234633180997, 0.0100138155060565, 0.0114190815800586]
IgnProb['046-01-03-L']=[0.000508781590063117, 0.00101256191269099, 0.0100138836975031, 0.0115814461696202]
IgnProb['046-01-04-L']=[0.00050866129882744, 0.00101235169916876, 0.0100138836975031, 0.0115698773836413]
IgnProb['062-01-01-L']=[0.000516144282959011, 0.00101235169916876, 0.0100138620131291, 0.0114396472822243]
IgnProb['043-01-02-L']=[0.000523907361472883, 0.00101252832811484, 0.0100138848346709, 0.0133282685238919]
IgnProb['043-02-02-L']=[0.000520386507843883, 0.00101255608114051, 0.010013883697503, 0]
IgnProb['General']=[0.000508858817833085, 0.00101259968232398, 0.0100138836975031, 0.011479454586573] #Condensate metering to be used for 'General'
# IgnProb['General']=[0.000502,0.001677,0.017112,0.017916]



IgnitionModel = "OLF"
p = []
#Set Ignition probability after reading the Discharge result
for e in lEvent:    
    # study,pv,pv_hole,hole,weather = e.Key.split("\\")
    pv,hole,weather = e.Key.split("\\")        
    key_hole = pv +"-"+ hole
    rr = ReleaseRate[key_hole] #shall be the value e.Discharge.ReleaseRate
    for k in IgnProb.keys():
        if k in pv:
            p = IgnProb[k]   
    
    if IgnitionModel == "OLF":
        # Adjustment for liquid mass fraction
        # rr *= (1-min(e.Discharge.LiquidMassFraction,0.95)) #Considering the additional 5% vapour only for the full liquid leak
        rr = rr * min(1 - e.Discharge.LiquidMassFraction + 0.05,1) #Considering the additional 5% vapour for all two-phase release 
        if p == []:
            p = IgnProb['General']
        if rr < 1:
            e.PImdIgn = 0.0005
            e.PDelIgn = p[0] - e.PImdIgn
        elif rr < 10:
            e.PImdIgn = 0.001
            e.PDelIgn = p[1] - e.PImdIgn        
        elif rr < 30.:
            e.PImdIgn = 0.01        
            e.PDelIgn = p[2] - e.PImdIgn        
        elif rr >= 30.:        
            e.PImdIgn = 0.01
            e.PDelIgn = p[3] - e.PImdIgn        


#Read TV Discharge data and append the info on each Event ... equip - hole - weather
for sfx in SFXFiles:
# for sfx in ['013-01-C2-G']:    
    tvExl = load_workbook(filename=TVDprefix + sfx+'.xlsx')
    print("Reading ...", sfx)
    sh = tvExl['Sheet1']
    new_scn = False
    #iterte for all row
    #read total row

    r = 1
    while r < sh.max_row:
        rv = sh.cell(r,1).value    
        #read the path at the row next toe 'Sceanrio ...'
        if rv != None:
            if "Scenario" in rv:
                #print (r, "in the loop")
                #print(rv)            
                r = r + 1
                path = sh.cell(r,1).value
                filename,study,pv,folder,hole = path.split("\\")                
                # path = path[SysLen:]
                path = study +"\\"+pv +"\\"+folder +"\\"+hole

                r = r + 2
                weather = sh.cell(r,1).value
                weather = weather[9:]
                #scn = path[20:]+"-"+weather[9:]
                scn = path +"\\"+weather
                r = r + 3
                t = sh.cell(r,1).value # read t=0 data
                mm = sh.cell(r,2).value
                rr = sh.cell(r,3).value
                m0 = mm
                r0 = rr
                TVD = [t, mm, rr ]            
                
                while t != None:
                    mm = sh.cell(r,2).value
                    rr = sh.cell(r,3).value
                    TVD = np.vstack([TVD, [t, mm, rr]])  
                    # if (("Flow line" in path) and ("NE_TLV" in path) and (t==3600)):
                    #     print([t, mm, rr])
                    #     print(TVD[-1,:])
                    tp = t
                    r = r + 1
                    t = sh.cell(r,1).value    
                                
                # if tp==3600:
                    # print(path, weather, TVD[-1,:])
                #print("Scenario {} ends at time {:8.2f} \n with the remaining mass {:10.2f}( \% of the inventory) and Release rate: {:8.2f})".\
                #    format(scn, tp, mm/m0*100, rr))
                #print("End of scenario {}".format(path+weather))        
                #print(r, rv)                


                #print out release rate at t=0, t=5min, t=10min, t=30min, t = 60min
                T_gate = [0, 300, 600, 1800, 3600]
                M_gate = [TVD[0,1], 0., 0., 0., 0.]
                RR_gate = [TVD[0,2], 0., 0., 0., 0.]
                #Tgi = 1
                for Tgi in range(1,5):
                    tp = 0
                    ti = 1
                    #if T_gate[Tgi] < Duration[scn]:
                    for t in TVD[1:,0]:                
                        if  tp < T_gate[Tgi] and T_gate[Tgi] <= t:
                            #t = TVD[ti,0]
                            M_gate[Tgi] = TVD[ti,1]
                            RR_gate[Tgi] = TVD[ti,2]
                            #print("ti={:5d} Mass {:12.2f} ({:12.2f}) \& Release Rate {:12.2f} ({:12.2f})at {:12.2f} read".format(ti,M_gate[Tgi],TVD[ti,1],RR_gate[Tgi],TVD[ti,2],t))
                        tp = t
                        ti = ti+1

                for e in lEvent:
                    if e.Path == path:
                        e.Discharge.Ts = T_gate
                        e.Discharge.Ms = M_gate
                        e.Discharge.RRs = RR_gate
                        e.TVDRead = True
                        e.TVD = TVD
                        print(e.Path,e.Weather, "TVD Read")
                



                if TVDPLot == True:
                    fig,ax1 = plt.subplots()                
                    Time = TVD[:,0]
                    Mass = TVD[:,1]
                    RelRate = TVD[:,2]

                    masscolor = 'tab:blue'
                    ax1.set_xlabel('Time [s]')
                    ax1.set_ylabel('Mass [kg]',color=masscolor)
                    ax1.plot(Time,Mass,color=masscolor)
                    ax1.set_ylim(bottom=0)
                    ax1.tick_params(axis='y',labelcolor=masscolor)
                    # ax1.xaxis.set_major_locator(plt.FixedLocator([300, 600, 1800, 3600]))
                    # ax1.set_xlim(left=0, right=3600)
                    
                    ax2 = ax1.twinx()
                    rrcolor = 'tab:red'
                    ax2.set_ylabel('Release Rate [kg/s]',color=rrcolor)
                    ax2.plot(Time,RelRate,color=rrcolor)
                    ax2.set_ylim(bottom=0)
                    ax2.tick_params(axis='y',labelcolor=rrcolor)
                    pngfilename = "TVD-"+pv+"_"+hole
                    plt.title(pngfilename)            
                    plt.show()
                    fn = slugify(pngfilename)
                    fig.savefig(".\\tvd_rev.B\\{}.png".format(fn))
                    plt.close()
                    

                #Copy the discharge pattern from 6.6/D to 1.2/F
                # i=0
                # while (not (lEvent[i].Path == path and lEvent[i].Weather == '1.2/F')):
                #     i = i + 1
                #     if i == len(lEvent):
                #         break
                # if i < len(lEvent):
                #     key = path + "\\" + '1.2/F'
                #     lEvent[i].Discharge.Ts = T_gate
                #     lEvent[i].Discharge.Ms = M_gate
                #     lEvent[i].Discharge.RRs = RR_gate
                #     lEvent[i].TVDRead = True
                #     lEvent[i].TVD = TVD
                #     #print(lEvent[i].Discharge.Ms)
                # else:
                #     print("Event corresponding to {} {} not found".format(path, '1.2/F'))                
        else:
            print("skip line {}".format(r))
            # continue
        r = r + 1

#End of reading Discharge files

#For an Event, find the relevant Leak and return its frequency
#r = 10
#key = lEvent[r].Path
#print("Path: {}  Frequency: {} Release Height: {}".format(key,Frequency[key],ReleaseHeight[key]))

#Reading Jet fire
#1. Find an Event that has the same study_folder_pv, scenario, & weather
# lEvent[##].JetFire = JetFiire
#for r in range(2,numLeak*len(weathers)+2):
for r in range(2,shJet.max_row+1):
    study_folder_pv = shJet.cell(r,1).value
    scenario = shJet.cell(r,2).value
    weather = shJet.cell(r,3).value
    path = study_folder_pv + "\\" + scenario
        
    study,pv,folder = study_folder_pv.split("\\")
    # holesizeindex = ['SM','ME','MA','LA'].index(scenario[0:2])

    key = pv + "\\" + scenario + "\\" + weather
    #Find the relevant event
    # for e in lEvent:
    #     if e.Key == key: 
    #         break
    
    i=0
    while not (lEvent[i].Path == path and lEvent[i].Weather == weather):
        i = i + 1
        if i == len(lEvent):
            break
    if i < len(lEvent):
        # key = path + "\\" + weather
        if weather in weathers:
            jfl = shJet.cell(r,10).value
            #J00[key] = jfl
            sep = shJet.cell(r,11).value
            JetSEP[key] = sep
            d04 = shJet.cell(r,15).value
            d12 = shJet.cell(r,16).value
            d37 = shJet.cell(r,17).value

            e = lEvent[i]
            jff = e.Frequency*e.PImdIgn/NumDirections            
            JetFrequency[key] = jff
            lEvent[i].JetFire = JetFire(jfl,sep,jff,d04,d12,d37)

            if e.Discharge.Ts[2] != 0:
                jf0 = e.JetFire.Length
                m0 = e.Discharge.RRs[0]
                jf0fit = jffit(m0)
                jfscale = e.JetFire.Length/jf0fit
                e.jfscale = jfscale
                lEvent[i].JetFire.JetLengths[0] = jf0
                for ti in range(1,5):
                    #t = a_event.Ts[i]
                    m = e.Discharge.RRs[ti]
                    jf = jffit(m)
                    lEvent[i].JetFire.JetLengths[ti] = jf*jfscale
                    #lEvent[i].JetFire.JetLengths[ti] = jf
                J00[key] = jf0
                J05[key] = lEvent[i].JetFire.JetLengths[1]
                J10[key] = lEvent[i].JetFire.JetLengths[2]
                J30[key] = lEvent[i].JetFire.JetLengths[3]
                J60[key] = lEvent[i].JetFire.JetLengths[4]

        else:
            print("{} Jet fire events read".format(k))
            break
    # k=k+1
#Read Jet, shJet


#shExplosion
ExpD1 = {}
ExpD2 = {}
ExpD3 = {}
ExpFreq = {}
#Read Early Pool Fire
for r in range(2,shExplosion.max_row+1):
    study_folder_pv = shExplosion.cell(r,1).value
    scenario = shExplosion.cell(r,2).value
    weather = shExplosion.cell(r,3).value
    # path = study_folder_pv + "\\" + scenario    
    # key = path + "\\" + weather    
    p1 = shExplosion.cell(r,9).value    
    p2 = shExplosion.cell(r,10).value    
    p3 = shExplosion.cell(r,11).value    
    d1 = shExplosion.cell(r,12).value    
    d2 = shExplosion.cell(r,13).value    
    d3 = shExplosion.cell(r,14).value    

    study,pv,folder = study_folder_pv.split("\\")
    key = pv + "\\" + scenario + "\\" + weather

    ExpD1[key] = d1
    ExpD2[key] = d2
    ExpD3[key] = d3
    
    for e in lEvent:
        if e.Key == key:
            # fexp = e.Frequency*(1-e.PImdIgn)*e.PDelIgn*e.PExp_Ign
            fexp = e.Frequency*e.PDelIgn*e.PExp_Ign
            ExpFreq[key] = fexp
            e.Explosion = Explosion(fexp,[p1,p2,p3],[d1,d2,d3])


#shDispersion
DistLFLFrac = {}
DistLFL = {}
DispFreq = {}
#Read Flammable Dispersion
for r in range(2,shDispersion.max_row+1):
    study_folder_pv = shDispersion.cell(r,1).value
    scenario = shDispersion.cell(r,2).value
    weather = shDispersion.cell(r,3).value
    path = study_folder_pv + "\\" + scenario    
    # key = path + "\\" + weather    
    # key = pv+ "\\" + hole + "\\" + weather    

    study,pv,folder = study_folder_pv.split("\\")
    key = pv + "\\" + scenario + "\\" + weather
    # 15,16,17
    dfmax = shDispersion.cell(r,15).value    
    dfmin = shDispersion.cell(r,16).value    
    wf = shDispersion.cell(r,17).value    
    # 19,20,21
    dmax = shDispersion.cell(r,19).value    
    dmin = shDispersion.cell(r,20).value    
    w = shDispersion.cell(r,21).value          

    ffmax = shDispersion.cell(r,27).value          
    ffw = shDispersion.cell(r,29).value          
    
    
    DistLFLFrac[key] = dfmax
    DistLFL[key] = dmax
    
    for e in lEvent:
        if e.Key == key:
            # fdisp = e.Frequency*(1-e.PImdIgn)*e.PDelIgn*(1-e.PExp_Ign)
            fdisp = e.Frequency*(1-e.PImdIgn - e.PDelIgn)
            DispFreq[key] = fdisp
            e.Dispersion = Dispersion(fdisp,dfmax,dfmin,wf,dmax,dmin,w,ffmax,ffw)

if 'Early Pool Fire' in cExl.sheetnames:
    shPool = cExl['Early Pool Fire']
    EPSEP = {}
    EPFreq = {}
    #Read Early Pool Fire
    for r in range(2,shPool.max_row+1):
        study_folder_pv = shPool.cell(r,1).value
        scenario = shPool.cell(r,2).value
        weather = shPool.cell(r,3).value
        path = study_folder_pv + "\\" + scenario    
        # key = path + "\\" + weather    
        study,pv,folder = study_folder_pv.split("\\")
        key = pv + "\\" + scenario + "\\" + weather

        sep = shPool.cell(r,11).value    
        EPSEP[key] = sep
        dia = shPool.cell(r,10).value    
        d04 = shPool.cell(r,15).value    
        d12 = shPool.cell(r,16).value    
        
        for e in lEvent:
            if e.Key == key:
                epff = e.Frequency*e.PImdIgn
                EPFreq[key] = epff
                e.EarlyPoolFire = EarlyPoolFire(dia,sep,epff,d04,d12)
                break



if 'Late Pool Fire' in cExl.sheetnames:
    shPool = cExl['Late Pool Fire']
    LPSEP = {}
    LPFreq = {}
    #Read Late Pool Fire
    for r in range(2,shPool.max_row+1):
        study_folder_pv = shPool.cell(r,1).value
        scenario = shPool.cell(r,2).value
        weather = shPool.cell(r,3).value
        # path = study_folder_pv + "\\" + scenario    
        # key = path + "\\" + weather    
        study,pv,folder = study_folder_pv.split("\\")
        key = pv + "\\" + scenario + "\\" + weather
        sep = shPool.cell(r,11).value    
        LPSEP[key] = sep
        dia = shPool.cell(r,10).value    
        d04 = shPool.cell(r,15).value    
        d12 = shPool.cell(r,16).value    
        for e in lEvent:
            if e.Key == key:
                # epff = e.Frequency*(1-e.PImdIgn)*e.PDelIgn
                epff = e.Frequency*e.PDelIgn
                LPFreq[key] = epff
                e.LatePoolFire = LatePoolFire(dia,sep,epff,d04,d12)


#shFireball
if 'Fireball' in cExl.sheetnames:
    shFireball = cExl['Fireball']
    DiaFireball = {}
    SEPFireball = {}
    FireballFreq = {}
    #Read Flammable Dispersion
    for r in range(2,shFireball.max_row+1):
        study_folder_pv = shFireball.cell(r,1).value
        scenario = shFireball.cell(r,2).value
        weather = shFireball.cell(r,3).value
        # path = study_folder_pv + "\\" + scenario    
        # key = path + "\\" + weather    
        study,pv,folder = study_folder_pv.split("\\")
        key = pv + "\\" + scenario + "\\" + weather
        
        dia = shFireball.cell(r,9).value    
        sep = shFireball.cell(r,10).value    
        
        DiaFireball[key] = dia
        SEPFireball[key] = sep
        
        for e in lEvent:
            if e.Key == key:
                ffireball = e.JetFire.Frequency
                FireballFreq[key] = ffireball
                e.Fireball = Fireball(ffireball,dia,sep)
for b in Bund:
    for e in lEvent:
        if b in e.Key:
            e.BundArea = Bund[b]

with open(element_dump_filename,'wb') as element_dump:
    dill.dump(lEvent,element_dump)
