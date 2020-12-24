
#2020. 10. 28 
#shET.cell(31,13).value = e.Frequency*(1-e.PDelIgn)*(1-P_GD_Fail) #Gas disperison #Wrong w.r.t. that e.PImdIgn is not subtracted
# #(1-e.PDelIgn) --> (1-e.PImdIgn-e.PDelIgn)
# Release rate to determine the detection probability ... 
# Option 1: Both to consider only the vapour
# Option 2: Fire detection using the total release rate and Gas detection using the vapour release rate?
# ll 93- 


from openpyxl import load_workbook
import os
import win32com.client
import time
import dill

NumDirections = 6

class Explosion:
    def __init__(self,fexp,op,d2op):
        self.Overpressures = op
        self.Distance2OPs = d2op
        self.Frequency = fexp        
    def __str__(self):
        a,b,c = self.Distance2OPs
        pa,pb,pc = self.Overpressures
        fmt = "{:.2e} | {:8.2f}barg to {:8.2f} | {:8.2f}bar to {:8.2f} | {:8.2f} barg to {:8.2f}".\
            format(self.Frequency, pa, a, pb, b, pc,c)        
        return fmt

# def EventTree_OLF2007(pv,hole,weather,lEvent,print2pdf=False):
print2pdf = True

# Area = 'ProcessArea'
# element_dump_filename = 'Bv08_c5_dump'
# icubeloc='SCE_CUBE_XYZ2_Process'

# element_dump_filename = 'Bv06_hull_dump'
# icubeloc='SCE_CUBE_XYZ2_HullDeck'

# element_dump_filename = 'Bv06_offloading_dump'
# icubeloc='SCE_CUBE_XYZ2_Offloading'

# element_dump_filename = 'Bv06_utility_dump'
# icubeloc='SCE_CUBE_XYZ2_Ulility'

# DUMP = {'Utility':'Bv06_utility_dump_2','ProcessArea':'Bv08_c5_dump','Hull':'Bv06_hull_dump','Offloading':'Bv06_offloading_dump'}
DUMP = {'ProcessArea':'Bv08_c5_dump'}
# DUMP = {'Hull':'Bv06_hull_dump','Offloading':'Bv06_offloading_dump'}
# DUMP = {'Utility':'Bv06_utility_dump_2'}


# c = 0
# for e in lEvent:
#     if ('020-04-01-L' in e.Key) and ('ME' in e.Hole) and ('14.5' in e.Weather):
#         break
#     c += 1
# print(c)
for d in DUMP:
    Area = d
    element_dump_filename = DUMP[d]

    with open(element_dump_filename,'rb') as element_dump:
        lEvent = dill.load(element_dump)
    ET_list = {}

    for e in lEvent:
    # for e in [lEvent[349]]:
        pv, hole, weather = e.Key.split("\\")
        hole = hole[:2]
        et_filename = 'C:\\Users\\seoshk\\LR\Energy - PRJ11100223773 - Documents\\6. Project Work place\\01 FRA\\PyExdCrv\\Rev.B\\et\\'+Area+'\\ET_'+pv+'_'+hole+'_'+weather+'.xlsx'
        path_to_pdf = et_filename[:-5]+".pdf"
        xlsfilename = 'ET_'+pv+'_'+hole+'_'+weather+'.xlsx'
        pdffilename = 'ET_'+pv+'_'+hole+'_'+weather+'.pdf'

        if (not (pv+hole+weather in ET_list.keys())) and (not (pdffilename in os.listdir(".\\et\\"+Area))):
            ET_list[pv+hole+weather] = True
            print(pv+hole+weather)
            import shutil
            #Failure of fire detector for immediate ignition
            P_FD_Fail = 0.0
            #Failure of gas detectin for delayed ignition
            P_GD_Fail = 0.0
            if xlsfilename in os.listdir(".\\et\\"):
                iExl=load_workbook(filename=et_filename)
            else:
                if (not ("020-03-01" in e.Key)) and ("-L" in e.Key):
                    shutil.copy('ET_template_NoBDV.xlsx',et_filename)
                else:
                    shutil.copy('ET_template.xlsx',et_filename)
                iExl=load_workbook(filename=et_filename)    
            
            shET = iExl['et']
            basic_freq_set = False
            for e in lEvent:
                if ((pv in e.Key) and (hole in e.Hole) and (weather == e.Weather)):
                    if e.Discharge.ReleaseRate <= 1:
                        P_FD_Fail = 0.1
                        P_GD_Fail = 0.1
                    elif e.Discharge.ReleaseRate <= 10:
                        P_FD_Fail = 0.05
                        P_GD_Fail = 0.05
                    elif e.Discharge.ReleaseRate > 10:
                        P_FD_Fail = 0.005
                        P_GD_Fail = 0.005
                    else:
                        print('something wrong in P_FD_Fail')
                    epv,ehole,eweather = e.Key.split("\\")
                    print("{:20s}{:10s} {:6s} {:8.2e} {:8.2e} {:8.2e} - Leak Freq {:8.2e} ".format(epv,ehole,eweather,e.Frequency,e.PESD,e.PBDV,e.Frequency/e.PESD/e.PBDV))
                    if basic_freq_set == False:
                        shET.cell(1,6).value = pv
                        shET.cell(1,7).value = hole
                        shET.cell(1,8).value = weather
                        shET.cell(30,2).value = e.Frequency/e.PESD/e.PBDV
                        shET.cell(31,2).value = e.Discharge.ReleaseRate
                        shET.cell(33,2).value = e.Discharge.ReleaseRate * min(1 - e.Discharge.LiquidMassFraction + 0.05,1)
                        shET.cell(20,6).value = e.PImdIgn
                        shET.cell(26,10).value = e.PDelIgn
                        shET.cell(15,7).value = 1-P_FD_Fail
                        shET.cell(42,7).value = 1-P_GD_Fail
                        # shET.cell(10,8).value = 1-0.01*numESDVs[pv]
                        shET.cell(10,8).value = e.PESD #PESD: p of successful ESD
                        # if "BN" in e.Hole:
                        #     shET.cell(9,9).value = 1
                        # else:
                        #     shET.cell(9,9).value = 1-0.005
                        shET.cell(9,9).value = e.PBDV #PBDV: p of successful BDV

                        basic_freq_set = True
                    else:
                        if abs(shET.cell(30,2).value - e.Frequency/e.PESD/e.PBDV)/(e.Frequency/e.PESD/e.PBDV) > 0.01:
                            print(pv, epv, ehole,'Basic frequency wrong',shET.cell(30,2).value, e.Frequency/e.PESD/e.PBDV, e.PESD, e.PBDV)
                            # break
                        # else:
                        #     print(epv,hole,'Basic frequency match')
                            
                        if shET.cell(20,6).value != e.PImdIgn:
                            print(epv, hole, 'ImdIgn Wrong')
                            # break
                        # else:
                        #     print(epv,hole,'ImdIgn match')
                            
                    if e.Explosion == None:
                        e.Explosion = Explosion(0.,0.,0.)
                    #Eithter fire detection for jet fire or gas detection for explosion has succedded
                    if ("EOBO" in ehole) or ("EOBN" in ehole):
                        shET.cell(9,13).value = e.JetFire.Frequency*NumDirections*(1-P_FD_Fail) #Jet fire frequency = LeakFreq x PEO x PBO x PImgIgn / NumDir
                        shET.cell(25,13).value = e.Explosion.Frequency*(1-P_GD_Fail) #VCE                
                        # shET.cell(28,13).value = e.Frequency*(1-e.PImdIgn)*(e.PDelIgn)*(1-e.PExp_Ign)*(1-P_GD_Fail) #Flash fire
                        # shET.cell(31,13).value = e.Frequency*(1-e.PImdIgn)*(1-e.PDelIgn)*(1-P_GD_Fail) #Gas disperison
                        shET.cell(28,13).value = e.Frequency*(e.PDelIgn)*(1-e.PExp_Ign)*(1-P_GD_Fail) #Flash fire
                        #shET.cell(31,13).value = e.Frequency*(1-e.PDelIgn)*(1-P_GD_Fail) #Gas disperison #Wrong w.r.t. that e.PImdIgn is not subtracted
                        shET.cell(31,13).value = e.Frequency*(1-e.PImdIgn-e.PDelIgn)*(1-P_GD_Fail) #Gas disperison
                        # print(pv,epv,ehole,shET.cell(9,13).value, shET.cell(25,13).value, shET.cell(31,13).value)
                    if "EOBX" in ehole:
                        shET.cell(12,13).value = e.JetFire.Frequency*NumDirections*(1-P_FD_Fail)
                        shET.cell(34,13).value = e.Explosion.Frequency*(1-P_GD_Fail)                
                        # shET.cell(37,13).value = e.Frequency*(1-e.PImdIgn)*(e.PDelIgn)*(1-e.PExp_Ign)*(1-P_GD_Fail) #Flash fire
                        # shET.cell(40,13).value = e.Frequency*(1-e.PImdIgn)*(1-e.PDelIgn)*(1-P_GD_Fail)
                        shET.cell(37,13).value = e.Frequency*(e.PDelIgn)*(1-e.PExp_Ign)*(1-P_GD_Fail) #Flash fire
                        shET.cell(40,13).value = e.Frequency*(1-e.PImdIgn-e.PDelIgn)*(1-P_GD_Fail)
                        # print(pv,epv,ehole,shET.cell(12,13).value, shET.cell(34,13).value, shET.cell(37,13).value)
                    if ("EXBO" in ehole):
                        shET.cell(15,13).value = e.JetFire.Frequency*NumDirections*(1-P_FD_Fail)
                        shET.cell(43,13).value = e.Explosion.Frequency*(1-P_GD_Fail)
                        # shET.cell(46,13).value = e.Frequency*(1-e.PImdIgn)*(e.PDelIgn)*(1-e.PExp_Ign)*(1-P_GD_Fail) #Flash fire
                        # shET.cell(49,13).value = e.Frequency*(1-e.PImdIgn)*(1-e.PDelIgn)*(1-P_GD_Fail)
                        shET.cell(46,13).value = e.Frequency*(e.PDelIgn)*(1-e.PExp_Ign)*(1-P_GD_Fail) #Flash fire                        
                        shET.cell(49,13).value = e.Frequency*(1-e.PImdIgn-e.PDelIgn)*(1-P_GD_Fail)
                        # print(pv,epv,ehole,shET.cell(15,13).value, shET.cell(43,13).value, shET.cell(49,13).value)
                    if ("EXBN" in ehole):
                        shET.cell(15,13).value = e.JetFire.Frequency*NumDirections*(1-P_FD_Fail)
                        shET.cell(43,13).value = e.Explosion.Frequency*(1-P_GD_Fail)
                        shET.cell(46,13).value = e.Frequency*(e.PDelIgn)*(1-e.PExp_Ign)*(1-P_GD_Fail) #Flash fire
                        shET.cell(49,13).value = e.Frequency*(1-e.PImdIgn-e.PDelIgn)*(1-P_GD_Fail)                    
                        #the following one line is added at July 30 2020
                        shET.cell(22,13).value = e.JetFire.Frequency*NumDirections/e.PESD/e.PBDV*(P_FD_Fail)                
                        shET.cell(61,13).value = e.Explosion.Frequency/e.PESD/e.PBDV*(P_GD_Fail)                    
                        shET.cell(64,13).value = e.Frequency/e.PESD/e.PBDV*(e.PDelIgn)*(1-e.PExp_Ign)*(P_GD_Fail) #Flash fire
                        shET.cell(67,13).value = e.Frequency/e.PESD/e.PBDV*(1-e.PImdIgn-e.PDelIgn)*(P_GD_Fail)
                    if ("EXBX" in ehole):
                        #Two cases should be distinguished; either detectors are successful or no
                        #Detection Success but EXBX
                        shET.cell(18,13).value = e.JetFire.Frequency*NumDirections*(1-P_FD_Fail)
                        shET.cell(52,13).value = e.Explosion.Frequency*(1-P_GD_Fail) #explosion
                        # shET.cell(55,13).value = e.Frequency*(1-e.PImdIgn)*(e.PDelIgn)*(1-e.PExp_Ign)*(1-P_GD_Fail) #Flash fire
                        # shET.cell(58,13).value = e.Frequency*(1-e.PImdIgn)*(1-e.PDelIgn)*(1-P_GD_Fail) #Dispersion only                
                        shET.cell(55,13).value = e.Frequency*(e.PDelIgn)*(1-e.PExp_Ign)*(1-P_GD_Fail) #Flash fire
                        shET.cell(58,13).value = e.Frequency*(1-e.PImdIgn-e.PDelIgn)*(1-P_GD_Fail) #Dispersion only                

                        #Detection failure & EXBX
                        shET.cell(22,13).value = e.JetFire.Frequency*NumDirections/e.PESD/e.PBDV*(P_FD_Fail)                
                        shET.cell(61,13).value = e.Explosion.Frequency/e.PESD/e.PBDV*(P_GD_Fail)
                        # shET.cell(64,13).value = e.Frequency/e.PESD/e.PBDV*(1-e.PImdIgn)*(e.PDelIgn)*(1-e.PExp_Ign)*(P_GD_Fail) #Flash fire
                        # shET.cell(67,13).value = e.Frequency/e.PESD/e.PBDV*(1-e.PImdIgn)*(1-e.PDelIgn)*(P_GD_Fail)
                        shET.cell(64,13).value = e.Frequency/e.PESD/e.PBDV*(e.PDelIgn)*(1-e.PExp_Ign)*(P_GD_Fail) #Flash fire
                        shET.cell(67,13).value = e.Frequency/e.PESD/e.PBDV*(1-e.PImdIgn-e.PDelIgn)*(P_GD_Fail)
                        
                        # print(pv,epv,ehole,shET.cell(22,13).value, shET.cell(61,13).value, shET.cell(66,13).value)

            iExl.save(et_filename)
            time.sleep(1)
            iExl.close()
            # print(et_filename)

    #If PDF-plotting is to be made separately, ...
    # for e in [lEvent[25]]:
    #     pv, hole, weather = e.Key.split("\\")
    #     hole = hole[:2]
    #     if pv+hole+weather in ET_list.keys():    
    #         et_filename = '.\\et\\ET_'+pv+'_'+hole+'_'+weather+'.xlsx'
            if print2pdf == True:        
                o = win32com.client.Dispatch("Excel.Application")
                o.Visible = False
                wb_path = et_filename
                wb = o.Workbooks.Open(wb_path)
                print_area = 'B1:M71'
                ws_index_list = [1] #say you want to print these sheets
                for index in ws_index_list:
                    #off-by-one so the user can start numbering the worksheets at 1
                    ws = wb.Worksheets[index - 1]
                    time.sleep(3)
                    ws.PageSetup.Zoom = False
                    ws.PageSetup.FitToPagesTall = 1
                    ws.PageSetup.FitToPagesWide = 1
                    ws.PageSetup.PrintArea = print_area           
                wb.WorkSheets(ws_index_list).Select()            
                
                
                wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
                wb.Close(False)
                o.Quit()
                time.sleep(1)