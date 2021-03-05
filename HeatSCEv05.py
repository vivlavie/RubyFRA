#Heat_SCE 
#v05 to provide coordiantes for the max heat point
#v04 to extract point-max value from each simulation, Pmax
#V03 at 2021.1.4 to reach new coordinates for flare lines
#To read heat on and around SCE's
#v02 updated only to read heat load on SCE's at 2020-11-03

from openpyxl import load_workbook
import math
import numpy as np
# import dill
import unicodedata
import re
import csv
import os
import time
import pandas as pd
from kfxtools import * #fit for Python 2.




# iHeat=load_workbook(filename='SCE_Heat_Template_20210104.xlsx')
# shHeatMax = iHeat['Max']
# shHeatAvg = iHeat['Avg']


#open an excel file for writing
iCoE=load_workbook(filename='SCE_CoE_20210305.xlsx',data_only=True)
shSCE = iCoE['CoE']
i = 2
SCE = {}
while shSCE.cell(i,3).value != None:
    tag = shSCE.cell(i,1).value
    sce_name = shSCE.cell(i,3).value
    key = tag + '_' + sce_name
    sce_type = shSCE.cell(i,2).value
    x = shSCE.cell(i,4).value
    y = shSCE.cell(i,5).value
    z = shSCE.cell(i,6).value
    Ori = shSCE.cell(i,9).value
    D = shSCE.cell(i,10).value
    H = shSCE.cell(i,11).value
    DZ = shSCE.cell(i,12).value
    Module = shSCE.cell(i,7).value
    Level = shSCE.cell(i,8).value        
    SCE[key] = [x,y,z,Ori,D,H,Module,Level,sce_name,sce_type,i,DZ]
    i += 1
nSCE = i - 2

fieldnum = 0

# P04_A_AP        : /projects/300341/Reliance/Fire/P04/J01
# P04_B_FS        : /projects/300341/Reliance/Fire/P04/J06
# S05_A_F           : /projects/300341/Reliance/Fire/S05/J09
# S05_B_F           : /projects/300341/Reliance/Fire/S05/J12
# S04_A_A          : /projects/300341/Reliance/Fire/S04/J13
# S03_A_P          : /projects/300341/Reliance/Fire/S03/J18
# S03_B_A          : /projects/300341/Reliance/Fire/P04/J19
NumRowsJetInfo = 5
Js = {}
Js['J01'] = ['P04_A_AP', 425.4]
# Js['J02'] = ['P04_A_FP', 407.2]
# Js['J03'] = ['P04_A_AS', 426.8]
# Js['J04'] = ['P04_A_FS', 452.5]
# Js['J05'] = ['P04_B_AS', 355.2]
# Js['J06'] = ['P04_B_FS', 345.8]
# Js['J07'] = ['S05_A_AS', 427.2]
# Js['J08'] = ['S05_A_AP', 452.5]
# # Js['J09'] = ['S05_A_F', 606.3]
# Js['J10'] = ['S05_A_F_10', 606.3] #00_J09 --> J09(5min) --> J10 (10 min) #For dose analysis, 00_J10 should direct not J09(23.5kg/sec) but 00_J09(61kg/sec)
# Js['J11'] = ['S05_B_AP', 426.8]
# Js['J12'] = ['S05_B_F', 351]
# Js['J13'] = ['S04_A_A', 334.1]
# Js['J14'] = ['S04_A_FP', 452.5]
# Js['J15'] = ['S04_A_FS', 305.2]
# Js['J16'] = ['S04_B_AP', 264.6]
# Js['J17'] = ['S04_B_F', 335.5]
# # Js['J18'] = ['S03_A_P', 163.8]
# # Js['J19'] = ['S03_B_A', 187]
# Js['J20'] = ['S03_B_F', 200.2]
# Js['J21'] = ['P02_B_FS', 130.4]
# Js['J22'] = ['P02_B_FP', 136.7]
# Js['J23'] = ['P05_A_P', 404.5]
# Js['J24'] = ['P05_A_S', 433.4]
# Js['J25'] = ['P05_B_A', 341.9]
# Js['J26'] = ['P05_B_F', 258.3]
# # Js['J27'] = ['P03_B_P', 159.1]
# Js['J28'] = ['P03_B_FS', 305]
# # Js['J29'] = ['KOD_B', 123.7]
# Js['P30'] = ['S05_A_AS', 452.5]
# Js['P31'] = ['S04_B_AP', 374.8]
# # Js['P31'] = ['S05_A_AP', 452.5]
# # Js['P32'] = ['S05_A_F', 452.5]



# Phase='00_'       
#or
# Phase=''       
PHASE = {'T0':'00_','T1':''}
pdP = {} #Point max
pdX = {} #X-coordiante for Point max
pdY = {} #Y-coordiante for Point max
pdZ = {} #Z-coordiante for Point max
pdM = {}
pdA = {}

for t in ['T0','T1']:
    Phase = PHASE[t]
    basefolder = "./"+Phase+"Rev.B/"

    FlareBox = {}
    pdPmax = pd.DataFrame(np.zeros((len(SCE),len(Js))),index=SCE.keys(), columns = Js)
    pdPmaxX = pd.DataFrame(np.zeros((len(SCE),len(Js))),index=SCE.keys(), columns = Js)
    pdPmaxY = pd.DataFrame(np.zeros((len(SCE),len(Js))),index=SCE.keys(), columns = Js)
    pdPmaxZ = pd.DataFrame(np.zeros((len(SCE),len(Js))),index=SCE.keys(), columns = Js)
    
    pdRmax = pd.DataFrame(np.zeros((len(SCE),len(Js))),index=SCE.keys(), columns = Js)
    pdRavg = pd.DataFrame(np.zeros((len(SCE),len(Js))),index=SCE.keys(), columns = Js)

    # sce_box = open('sce_box_flare_max_'+t+'.kfx','w')
    for jk in Js.keys():
    # for jk in ['J06', 'J09','J13','J13']:

        colid = int(jk[-2:])+12
        fdr = Js[jk][0][:3]
        # fn = basefolder + fdr + "/" + jk + "/" + jk+"_rad_exit.r3d"        
        fn = basefolder + "/" + Phase + jk + "_rad_exit.r3d"        
        
        if (os.path.exists(fn) == False):
            print(fn + " does not exist")    
        else:    
            #Rad radiation
            T = readr3d(fn)            
            fieldname = T.names[fieldnum]
            print(fieldname)
            print(Js[jk][0],fn,fieldname)

            for key in SCE.keys():        
            # for key in ['045-ES-001_FUEL GAS HEATER']:        
                # if SCE[k][6] in Js[jk]:
                s = SCE[key]
                x = s[0]
                y = s[1]
                z = s[2]
                Ori = s[3]
                D = s[4]
                H = s[5]
                sce_name = s[8]
                rowid = s[10]
                rmax = ravg = 0.
                # if ("020" in k) and ("SEP" in k):
                if True:
                    if Ori == 'X':
                        dx = H/2.
                        dy = D/2.
                        dz = dy
                    elif Ori == 'Y':
                        dy = H/2.
                        dx = D/2.
                        dz = dx
                    elif Ori == 'Z':
                        dz = H/2.
                        dy = D/2.
                        dx = dy
                    elif Ori == 'C':
                        dx = dy = dz = D/2.                
                    elif Ori == 'B':
                        dx = D/2.
                        dy = H/2.
                        dz = s[-1]/2. #s[-1] = DZ
                    elif Ori in ['XP','XN']:
                        dz = H/2./1.3
                        # z -= 0.8*dz
                        dy = D/2./1.3
                        dx = 0
                    elif Ori in ['YP','YN']:
                        dz = H/2./1.3
                        # z -= 0.8*dz
                        dx = D/2./1.3
                        dy = 0
                    elif Ori == None:
                        print('Orientation is not read')
                    else:
                        print('Orientation wrong input', Ori, k, s)
                    
                    Xs = [x - 1.3*dx, x, x + 1.3*dx]
                    Ys = [y - 1.3*dy, y, y + 1.3*dy]
                    Zs = [z - 1.3*dz, z, z + 1.3*dz]
                    R = np.zeros((3,3,3))
                    if Ori in ['X','Y','Z','C','B']:
                        # r = 0.
                        # rmax = 0.
                        # rsum = 0.
                        # ncount = 0
                        # ravg = 0.
                        C = []
                        P = []
                        for i,xi in enumerate(Xs):
                            for j,yi in enumerate(Ys):
                                if Ori != "C":
                                    for k,zi in enumerate(Zs):
                                        if not ((xi == x) and (yi == y) and (zi == z)):
                                            r = T.point_value(xi,yi,zi,fieldnum)                                
                                            R[i,j,k] = r                                        
                                            C.append([xi,yi,zi])
                                            P.append(r)
                                else: #for the crane only!                    
                                    for zi in [z]:
                                        if not ((xi == x) and (yi == y) and (zi == z)):
                                            r = T.point_value(xi,yi,zi,fieldnum)
                                            R[i,j,1] = r                                        
                                            C.append([xi,yi,zi])
                                            P.append(r)
                        Pmax = R.max()
                        Pmax2 = max(P)
                        if Pmax != Pmax2:
                            print("something wrong", R,P,C)
                        Cmax = C[argmax(P)]

                        Xmax = max(R[0,:,:].mean(),R[2,:,:].mean())
                        Ymax = max(R[:,0,:].mean(),R[:,2,:].mean())
                        Zmax = max(R[:,:,0].mean(),R[:,:,2].mean())

                        rmax = max(Xmax,Ymax,Zmax)
                        ravg = R.mean()
                            
                        # else:
                            # print(k,"something wrong if the following two numbers are not zero", ncount, rsum) 
                    elif Ori in ['XP','XN','YP','YN']:
                        if Ori in ['XP','XN']:
                            if Ori == 'XP':
                                xi = Xs[1] + 0.1
                                i = 2
                            elif Ori == 'XN':
                                xi = Xs[1] - 0.1
                                i = 0
                            for j,yi in enumerate(Ys):                            
                                for k,zi in enumerate(Zs):                                
                                    r = T.point_value(xi,yi,zi,fieldnum)
                                    R[i,j,k] = r                                        
                            ravg = max(R[0,:,:].mean(),R[2,:,:].mean())                        
                            rmax = max(R[0,:,:].max(),R[2,:,:].max())                        
                        if Ori in ['YP','YN']:
                            if Ori == 'YP':
                                yi = Ys[1] + 0.1
                                j = 2
                            elif Ori == 'YN':
                                yi = Ys[1] - 0.1
                                j = 0
                            for i,xi in enumerate(Xs):                            
                                for k,zi in enumerate(Zs):                                
                                    r = T.point_value(xi,yi,zi,fieldnum)
                                    R[i,j,k] = r                                        
                            ravg = max(R[:,0,:].mean(),R[:,2,:].mean())                        
                            rmax = max(R[:,0,:].max(),R[:,2,:].max())                        
                            Pmax = R.max()

                    else:
                        print("Orientation is wrong!",s)
                    
                    if (rmax == np.nan) or (rmax == np.NaN):
                        rmax = ravg = 0.
                    
                    # if (rmax*0.001 > 100):
                    # if (rmax*0.001 > 100) and (key.find("Flare_Line") > 0):
                    # if (key.find("GDU_Package") > 0):
                    # if ('P04' in s[6]):
                    #     sce_box.write("PART: {:15s} {:s} {:6.1f} {:6.1f} \nBOX: {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f} {:8.1f}\n".format(key,jk, rmax/1000,ravg/1000,(Xs[0])*1000,(Ys[0])*1000,(Zs[0])*1000,(Xs[2]-Xs[0])*1000,(Ys[2]-Ys[0])*1000,(Zs[2]-Zs[0])*1000))
                    
                    #Print out for 'flare'
                    # if ("Flare" in key) and (ravg > 50000):
                    #     if not (key in FlareBox.keys()):
                    #         FlareBox[key] = [rmax, ravg]
                    #     else:
                    #         FlareBox[key] = [max(FlareBox[key][0], rmax), max(FlareBox[key][1],ravg)]                        
                    
                    # shHeatMax.cell(rowid+NumRowsJetInfo,colid).value = "{:8.2f}".format(rmax/1000)
                    # shHeatAvg.cell(rowid+NumRowsJetInfo,colid).value = "{:8.2f}".format(ravg/1000)
                    # shHeatMax.cell(rowid+NumRowsJetInfo,colid).value = rmax/1000
                    # shHeatAvg.cell(rowid+NumRowsJetInfo,colid).value = ravg/1000

                    pdPmax.loc[key,jk] = Pmax/1000
                    pdPmaxX.loc[key,jk] = Cmax[0]
                    pdPmaxY.loc[key,jk] = Cmax[1]
                    pdPmaxZ.loc[key,jk] = Cmax[2]
                    # pdCmax.loc[key,jk] = Cmax#coordiates where the point max is read
                    pdRmax.loc[key,jk] = rmax/1000
                    pdRavg.loc[key,jk] = ravg/1000
                

    # iHeat.save('SCE_Heat_'+t+'_'+time.strftime("%Y%m%d-%H%M%S")+'.xlsx')
    # sce_box.close()
    pdP[t] = pdPmax
    pdX[t] = pdPmaxX
    pdY[t] = pdPmaxY
    pdZ[t] = pdPmaxZ
    
    # pdC[t] = pdCmax
    pdM[t] = pdRmax
    pdA[t] = pdRavg
      
#Choose which heat load is used for evaluating dose
idmax = pdM['T1'].idxmax(axis=1)
maxsces = idmax.index
pdPMax = pd.DataFrame(np.zeros((len(SCE),7)),index=SCE.keys(),columns=['Fire','Module','Deck','T0','T1','Dur','Dose'])
pdMax = pd.DataFrame(np.zeros((len(SCE),7)),index=SCE.keys(),columns=['Fire','Module','Deck','T0','T1','Dur','Dose'])
pdAvg = pd.DataFrame(np.zeros((len(SCE),7)),index=SCE.keys(),columns=['Fire','Module','Deck','T0','T1','Dur','Dose'])
for sce in maxsces:
    jk = idmax.loc[sce]    
    # print(sce,jk)
    if (sce in pdM['T1'].index):
        #Max
        pdPMax.loc[sce,'Fire'] = jk
        m0 = pdP['T0'].loc[sce,jk]
        pdPMax.loc[sce,'T0'] = m0
        m1 = pdP['T1'].loc[sce,jk]
        pdPMax.loc[sce,'T1'] = m1
        t1 = Js[jk][1]
        pdPMax.loc[sce,'Dur'] = t1
        pdPMax.loc[sce,'Module'] = SCE[sce][6]
        pdPMax.loc[sce,'Deck'] = SCE[sce][7]
        #No distinguishing depedning on which is larger m0 & m1
        pdPMax.loc[sce,'Dose'] = m0*120 + 0.5*(m0+m1)*(t1-120)
        
        #Max
        pdMax.loc[sce,'Fire'] = jk
        m0 = pdM['T0'].loc[sce,jk]
        pdMax.loc[sce,'T0'] = m0
        m1 = pdM['T1'].loc[sce,jk]
        pdMax.loc[sce,'T1'] = m1
        t1 = Js[jk][1]
        pdMax.loc[sce,'Dur'] = t1
        pdMax.loc[sce,'Module'] = SCE[sce][6]
        pdMax.loc[sce,'Deck'] = SCE[sce][7]
        #No distinguishing depedning on which is larger m0 & m1
        pdMax.loc[sce,'Dose'] = m0*120 + 0.5*(m0+m1)*(t1-120)
        # if m0 > m1:
        #     pdMax.loc[sce,'Dose'] = m0*120 + 0.5*(m0+m1)*(t1-120)
        # else:
        #     pdMax.loc[sce,'Dose'] = 0.5*(m0+m1)*120 + m1*(t1-120)
        
        #Avg
        pdAvg.loc[sce,'Fire'] = jk
        a0 = pdA['T0'].loc[sce,jk]
        pdAvg.loc[sce,'T0'] = a0
        a1 = pdA['T1'].loc[sce,jk]
        pdAvg.loc[sce,'T1'] = a1
        t1 = Js[jk][1]
        pdAvg.loc[sce,'Dur'] = t1
        pdAvg.loc[sce,'Module'] = SCE[sce][6]
        pdAvg.loc[sce,'Deck'] = SCE[sce][7]
        pdAvg.loc[sce,'Dose'] = a0*120 + 0.5*(a0+a1)*(t1-120)
        # if a0 > a1:
        #     pdAvg.loc[sce,'Dose'] = a0*120 + 0.5*(a0+a1)*(t1-120)
        # else:
        #     pdAvg.loc[sce,'Dose'] = 0.5*(a0+a1)*120 + a1*(t1-120)
    else:
        print('sce ', sce,' is not in index or fire scenario is ', jk)

# print("{:80s}{:10s}{:>20s}{:>20s}".format("SCE","Fire","Max Dose[MJ/m2]","Avg Dose[MJ/m2]"))
# for sce in maxsces:
#     print("{:80s}{:10s}{:20.1f}{:20.1f}".format(sce,pdMax.loc[sce,'Fire'],pdMax.loc[sce,'Dose']*0.001,pdAvg.loc[sce,'Dose']*0.001))
pdAll = pd.merge(pdPMax,pdMax,left_index=True,right_index=True,suffixes=["P","M"])   
pdAll = pd.merge(pdAll,pdAvg,left_index=True,right_index=True,suffixes=["","A"])   

with pd.ExcelWriter('VSA'+time.strftime("%Y%m%d-%H%M%S")+'.xlsx') as writer:  
    pdP['T0'].to_excel(writer, sheet_name='PointMaxT0')
    pdP['T1'].to_excel(writer, sheet_name='PointMaxT1')
    pdX['T0'].to_excel(writer, sheet_name='X_T0')
    pdX['T1'].to_excel(writer, sheet_name='X_T1')
    pdY['T0'].to_excel(writer, sheet_name='Y_T0')
    pdY['T1'].to_excel(writer, sheet_name='Y_T1')
    pdZ['T0'].to_excel(writer, sheet_name='Z_T0')
    pdZ['T1'].to_excel(writer, sheet_name='Z_T1')
    
    pdM['T0'].to_excel(writer, sheet_name='MaxT0')
    pdM['T1'].to_excel(writer, sheet_name='MaxT1')
    pdA['T0'].to_excel(writer, sheet_name='AvgT0')
    pdA['T1'].to_excel(writer, sheet_name='AvgT1')
    pdPMax.to_excel(writer, sheet_name='PointMaxDose')
    pdMax.to_excel(writer, sheet_name='MaxDose')
    pdAvg.to_excel(writer, sheet_name='AvgDose')
    pdAll.to_excel(writer, sheet_name='Dose')    
    

# pdMt1 = pdM['T1']
# is_Flareline = pdMt1.index.str.contains('Flare_Line')